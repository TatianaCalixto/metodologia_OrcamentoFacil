"""
Adiciona Sprints 17, 18 e 19 (recomendacoes da analise PO/UX/UI) na planilha
existente, preservando todo o tracking ja preenchido pelo chat executor.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


XLSX_PATH = r"d:\Projetos\OrcamentoFamiliar\orcafacil_planejamento_sprints.xlsx"


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
SPRINT_FILL = PatternFill("solid", fgColor="D9E1F2")
NEW_HIGHLIGHT = PatternFill("solid", fgColor="FFF2CC")  # destaque amarelo claro

STATUS_OPTIONS = [
    "Pendente", "Em andamento", "Bloqueada", "Em teste", "Concluída", "Cancelada"
]
PRIORIDADE_OPTIONS = ["Alta", "Média", "Baixa"]
TIPO_OPTIONS = ["Backend", "Frontend", "Mobile", "Infra", "Banco", "Teste", "Documentação", "Deploy"]


# =========================================================
# Conteudo das novas sprints
# =========================================================

FASE_TAG = "[FASE 2]"
FASE_OBS = ("FASE 2 do projeto — Sprints 0 a 16 = FASE 1 (MVP). "
            "Análise realizada em 2026-05-24 pelos papéis: Product Owner, "
            "UX Designer e UI Designer. Documento-base: "
            "analise_po_ux_ui_orcafacil.md.")

NOVAS_SPRINTS = [
    {
        "numero": 17,
        "nome": f"{FASE_TAG} Sprint 17 — Gestão de Contas (mobile + painel)",
        "objetivo": "Fechar a lacuna crítica P1 da análise PO/UX/UI: hoje não existe "
                    "tela para criar/editar conta em lugar nenhum, embora o backend "
                    "tenha o CRUD completo. Sem conta, o usuário novo não consegue "
                    "lançar transação.",
        "entregaveis": "Tela de Contas no mobile (lista + formulário), item 'Contas' "
                       "no drawer com a reorganização das seções, e página de Contas "
                       "no painel Streamlit. Bateria de regressão de saldo continua "
                       "verde após as mudanças.",
    },
    {
        "numero": 18,
        "nome": f"{FASE_TAG} Sprint 18 — Painel web com cadastro completo",
        "objetivo": "Fechar as lacunas P2 e P5: o painel hoje só lê e exporta. Usuário "
                    "desktop precisa abrir o celular para qualquer escrita. Estender "
                    "o painel para cadastro/edição/exclusão de transação e importação "
                    "CSV, mantendo paridade funcional com o mobile.",
        "entregaveis": "Páginas 'Nova Transação', 'Importar CSV' e ações de "
                       "edição/exclusão na lista de Transações do painel; "
                       "customização visual básica (config.toml + logo).",
    },
    {
        "numero": 19,
        "nome": f"{FASE_TAG} Sprint 19 — Polimento UX/UI",
        "objetivo": "Resolver problemas de fricção e consistência visual identificados "
                    "(P3, P4, P6, P7, P8, P9): atalho rápido para lançar transação, "
                    "onboarding pós-cadastro, design tokens, tema escuro real, "
                    "feedback pós-ação.",
        "entregaveis": "FAB no Dashboard, onboarding com 'Crie sua primeira conta', "
                       "tokens de tema, dark mode funcional, drawer com item ativo "
                       "destacado, snackbar de saldo pós-transação, picker de ícone "
                       "nas categorias.",
    },
]


# (sprint, id, tipo, titulo, descricao, criterios, deps, testes, prioridade)
NOVAS_TAREFAS = [
    # ---------- Sprint 17 — Gestão de Contas ----------
    (17, "S17-T01", "Mobile",
     "Tela 'Contas' (lista) no mobile",
     "Criar lib/features/accounts/ com data layer (api + models) consumindo "
     "GET/DELETE /accounts. Tela de lista com cards (nome, tipo, saldo atual, "
     "ativa/inativa), filtro por ativa, FAB '+' para nova conta, swipe ou "
     "menu para excluir com diálogo de confirmação. Estados loading/empty/error.",
     "Lista exibe contas do usuário autenticado; filtro de ativa/inativa funciona; "
     "excluir solicita confirmação e atualiza a lista; estados visuais corretos.",
     "—",
     "Widget tests: estado loading; estado vazio; estado com dados; confirmação "
     "de exclusão com mock do Dio.",
     "Alta"),
    (17, "S17-T02", "Mobile",
     "Tela 'Formulário de Conta' no mobile (criar/editar)",
     "Tela única que cria ou edita conta conforme parâmetro. Campos: nome, tipo "
     "(enum: corrente, poupança, dinheiro, cartão), saldo inicial (apenas na "
     "criação), switch ativa. Validações: nome obrigatório, saldo numérico. "
     "Rotas: /accounts/new e /accounts/:id/edit. Após submit volta para /accounts.",
     "Criar conta dispara POST /accounts; editar dispara PATCH; saldo inicial "
     "não é editável após criação; snackbar de sucesso ao concluir; erros do "
     "backend exibidos amigavelmente.",
     "S17-T01",
     "Widget tests do formulário em modo criar e editar (com mock); validação "
     "de campos obrigatórios; teste de submit feliz e teste de erro do backend.",
     "Alta"),
    (17, "S17-T03", "Mobile",
     "Item 'Contas' no drawer + reorganização das seções",
     "Adicionar item 'Contas' no drawer (ícone account_balance) entre 'Transações' "
     "e 'Categorias'. Reorganizar drawer em duas seções com Divider: "
     "'Lançamentos' (Dashboard, Transações, Contas, Categorias, Orçamentos, "
     "Metas, Importar CSV) e 'Conta' (Perfil, Configurações, Sair).",
     "Drawer mostra 'Contas' navegando para /accounts; sair (logout) está "
     "agora no drawer (não só no AppBar); seções visíveis com separador.",
     "S17-T01",
     "Widget test do drawer: existência do item 'Contas'; clique navega para "
     "/accounts; logout no drawer limpa token e volta para login.",
     "Alta"),
    (17, "S17-T04", "Frontend",
     "Página 'Contas' no painel Streamlit (CRUD completo)",
     "Criar panel/pages/3_Contas.py. Lista de contas em st.dataframe; "
     "expander/form para criar nova conta (nome, tipo, saldo inicial, ativa); "
     "edição via st.form com seleção da conta; exclusão com st.button + "
     "confirmação (st.dialog ou checkbox 'confirmo').",
     "Página acessível após login; CRUD completo funciona contra a API; "
     "operações refletem na listagem após rerun; erros exibidos via st.error.",
     "—",
     "Testes (panel/tests/) com mock da camada api.py cobrindo create, update "
     "e delete; smoke test manual documentado no README do panel.",
     "Alta"),

    # ---------- Sprint 18 — Painel web com cadastro completo ----------
    (18, "S18-T01", "Frontend",
     "Página 'Nova Transação' no painel",
     "Criar panel/pages/4_Nova_Transacao.py com formulário: tipo (income/expense), "
     "valor, data (date_input), conta (selectbox carregado via api), categoria "
     "(selectbox filtrado pelo tipo), descrição (text_input), forma de pagamento "
     "(selectbox opcional), recorrente (checkbox). Submit chama POST /transactions.",
     "Formulário submete corretamente; categorias filtram por tipo selecionado; "
     "validações de campos obrigatórios; snackbar (st.success) ao salvar; "
     "limpar formulário após salvar com sucesso.",
     "S17-T04",
     "Teste em panel/tests/ cobrindo: payload correto; filtro de categorias por "
     "tipo; tratamento de erro 422 do backend.",
     "Alta"),
    (18, "S18-T02", "Frontend",
     "Edição de transação na lista do painel",
     "Em panel/pages/2_Transacoes.py, adicionar coluna de ação com botão 'Editar' "
     "que abre formulário (st.dialog ou expander) pré-preenchido. Submit dispara "
     "PATCH /transactions/{id} e faz rerun para atualizar a lista.",
     "Edição funciona contra o backend; lista atualizada após edição; cache "
     "invalidado quando necessário.",
     "S18-T01",
     "Teste com mock cobrindo edição feliz e edição com erro do backend.",
     "Alta"),
    (18, "S18-T03", "Frontend",
     "Exclusão de transação no painel com confirmação",
     "Adicionar ação 'Excluir' por transação na lista. Pedir confirmação "
     "(st.dialog 'Tem certeza? Esta ação não pode ser desfeita'). Submit dispara "
     "DELETE /transactions/{id} e atualiza a lista.",
     "Exclusão pede confirmação explícita; lista atualizada após confirmar; "
     "snackbar de sucesso.",
     "S18-T02",
     "Teste com mock cobrindo confirmação aceita, confirmação cancelada e "
     "erro do backend.",
     "Alta"),
    (18, "S18-T04", "Frontend",
     "Página 'Importar CSV' no painel",
     "Criar panel/pages/5_Importar_CSV.py com st.file_uploader para CSV, "
     "selectbox de conta alvo, botão 'Importar'. Submit chama POST /imports/csv "
     "(multipart) e exibe resumo: total processadas, sucesso, erros (lista).",
     "Upload válido importa; arquivo inválido exibe erro; resumo de erros é "
     "compreensível; saldo da conta reflete o import (verificável na aba "
     "Relatórios).",
     "S17-T04",
     "Teste em panel/tests/ com mock do upload cobrindo: CSV válido; "
     "CSV inválido; resumo de erros parciais.",
     "Alta"),
    (18, "S18-T05", "Frontend",
     "Customização visual básica do painel",
     "Criar panel/.streamlit/config.toml com paleta consistente com o app mobile "
     "(primaryColor, backgroundColor, secondaryBackgroundColor, textColor). "
     "Adicionar logo simples (texto estilizado + ícone) no topo do app.py e "
     "em cada página via componente comum.",
     "Painel exibe paleta personalizada (não mais o tema default do Streamlit); "
     "logo presente em todas as páginas; sem regressão funcional.",
     "S18-T04",
     "Verificação visual manual documentada com screenshots no panel/README.md.",
     "Média"),

    # ---------- Sprint 19 — Polimento UX/UI ----------
    (19, "S19-T01", "Mobile",
     "FAB '+' no Dashboard para Nova Transação",
     "Adicionar FloatingActionButton no DashboardScreen que navega diretamente "
     "para /transactions/new. Ícone Icons.add, key 'btn-dashboard-new-tx'.",
     "FAB visível no Dashboard; clique navega para a tela de Nova Transação; "
     "ação primária acessível em 1 toque a partir do Dashboard.",
     "—",
     "Widget test: FAB presente no Dashboard; clique aciona navegação para "
     "/transactions/new (mock de GoRouter).",
     "Alta"),
    (19, "S19-T02", "Mobile",
     "Onboarding pós-cadastro: 'Crie sua primeira conta'",
     "Após RegisterScreen ou Login (quando o usuário não tem nenhuma conta), "
     "exibir modal/tela obrigatória 'Crie sua primeira conta' com link para "
     "/accounts/new. Permitir 'Pular por agora' com aviso 'Sem conta você não "
     "consegue lançar transações'. Verificação via GET /accounts: se lista "
     "vazia, mostra onboarding uma vez.",
     "Novo usuário vê o onboarding na primeira sessão; após criar conta, não "
     "vê de novo; 'Pular' funciona e mostra aviso.",
     "S17-T02",
     "Widget tests: usuário sem conta vê onboarding; usuário com conta não vê; "
     "'Pular' fecha o modal mas avisa.",
     "Alta"),
    (19, "S19-T03", "Mobile",
     "Tokens de tema em lib/core/theme/",
     "Criar lib/core/theme/app_colors.dart (cores semânticas: primary, success, "
     "warning, danger, surface, onSurface, etc.), app_text_styles.dart "
     "(tipografia escalonada), app_spacing.dart (escala 4/8/12/16/24/32). "
     "Refatorar telas existentes substituindo Colors.green/red/blueGrey "
     "hardcoded por referências do tema.",
     "Todas as cores hardcoded substituídas; tema configurado em MaterialApp "
     "consome os tokens; nenhuma tela visualmente diferente do estado anterior.",
     "—",
     "flutter analyze limpo após refatoração; widget tests pré-existentes "
     "continuam verdes; golden tests opcionais comparando antes/depois.",
     "Alta"),
    (19, "S19-T04", "Mobile",
     "Tema escuro funcional",
     "Definir AppColors.darkScheme (paleta de contraste validada com WCAG AA). "
     "MaterialApp consome themeMode do SettingsController. Garantir que todos "
     "os widgets usam tokens (não cores fixas).",
     "Alternar tema em Configurações muda visualmente o app inteiro; sem "
     "regressão visual nas telas; contraste mínimo 4.5:1 em texto normal.",
     "S19-T03",
     "Golden tests dos 2 modos (claro/escuro) nas 5 telas principais: "
     "Dashboard, Transações, Contas, Categorias, Configurações.",
     "Alta"),
    (19, "S19-T05", "Mobile",
     "Highlight do item ativo no NavigationDrawer",
     "Substituir Drawer + ListView por NavigationDrawer/NavigationDrawerDestination "
     "do Material 3, com selectedIndex sincronizado pela rota atual via GoRouter.",
     "Item correspondente à rota atual aparece destacado (cor secundária) ao "
     "abrir o drawer; navegação continua funcionando.",
     "S17-T03",
     "Widget test: drawer aberto em /transactions destaca 'Transações'; em "
     "/budgets destaca 'Orçamentos'.",
     "Média"),
    (19, "S19-T06", "Mobile",
     "Snackbar de saldo pós-transação",
     "Após criar ou editar transação com sucesso, snackbar mostra 'Transação "
     "salva. Saldo de {conta}: R$ X.XX'. Buscar saldo via GET /accounts/{id} "
     "ou pela conta no controller já carregado.",
     "Snackbar exibe novo saldo da conta afetada; valor formatado em BRL; "
     "no caso de edição que mudou a conta, mostra a conta nova.",
     "S19-T03",
     "Widget test cobrindo: snackbar pós-criação; snackbar pós-edição; "
     "formato BRL correto.",
     "Média"),
    (19, "S19-T07", "Mobile",
     "Picker de ícone na tela de Categorias",
     "Em CategoryFormScreen, substituir input de texto livre do campo 'icon' por "
     "grid picker com ~20 ícones do Material (Icons.restaurant, "
     "Icons.directions_car, Icons.local_hospital, Icons.school, Icons.home, "
     "Icons.movie, Icons.subscriptions, etc.). Persistir o nome do ícone "
     "(string) no backend para compatibilidade.",
     "Picker exibe grade de ícones; seleção fica destacada; valor salvo no "
     "backend é a string do ícone; tela de Categorias exibe o ícone correto.",
     "—",
     "Widget test: picker abre; seleção muda o estado; submit envia a string "
     "correta; teste do mapeamento string -> IconData.",
     "Baixa"),
]


# Itens de Plano de Testes adicionais
# (sprint_alvo, modulo, tipo_teste, descricao)
NOVOS_TESTES = [
    (17, "accounts (mobile)", "Widget", "Tela Contas nos 3 estados (loading/vazio/com dados)"),
    (17, "accounts (mobile)", "Widget", "Formulário de conta criar/editar com mock"),
    (17, "accounts (mobile)", "Integração", "E2E: criar conta -> lançar transação -> saldo correto"),
    (17, "accounts (panel)", "Unitário", "CRUD de conta no painel com mock da api"),
    (17, "regressão", "Regressão", "test_balance_regression_full_flow continua verde após S17"),

    (18, "panel transactions", "Unitário", "Nova Transação: payload correto + filtro de categorias por tipo"),
    (18, "panel transactions", "Unitário", "Edição via lista do painel (PATCH mock)"),
    (18, "panel transactions", "Unitário", "Exclusão com confirmação (DELETE mock)"),
    (18, "panel imports", "Unitário", "Upload de CSV no painel (válido / inválido / erros parciais)"),
    (18, "regressão", "Smoke", "E2E painel: login -> criar conta -> criar transação -> ver no relatório"),
    (18, "regressão", "Regressão", "test_balance_regression_full_flow continua verde após S18"),

    (19, "mobile dashboard", "Widget", "FAB no Dashboard navega para /transactions/new"),
    (19, "mobile onboarding", "Widget", "Usuário sem conta vê modal; com conta não vê; Pular avisa"),
    (19, "mobile theme", "Golden", "5 telas principais nos modos claro e escuro"),
    (19, "mobile drawer", "Widget", "Item ativo do drawer reflete a rota atual"),
    (19, "mobile transactions", "Widget", "Snackbar pós-transação exibe novo saldo formatado BRL"),
    (19, "mobile categories", "Widget", "Picker de ícone abre, seleciona e persiste corretamente"),
]


# =========================================================
# Helpers
# =========================================================

def style_body(cell, highlight=False):
    cell.alignment = WRAP
    cell.border = BORDER
    if highlight:
        cell.fill = NEW_HIGHLIGHT


def append_sprints_overview(ws):
    """Anexa as novas sprints na aba Visao Geral."""
    start_row = ws.max_row + 1
    for i, s in enumerate(NOVAS_SPRINTS):
        r = start_row + i
        ws.cell(row=r, column=1, value=s["numero"])
        ws.cell(row=r, column=2, value=s["nome"])
        ws.cell(row=r, column=3, value=s["objetivo"])
        ws.cell(row=r, column=4, value=s["entregaveis"])
        ws.cell(row=r, column=5, value="Pendente")
        ws.cell(row=r, column=6, value="")
        ws.cell(row=r, column=7, value="")
        ws.cell(row=r, column=8, value=FASE_OBS)
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            style_body(cell, highlight=True)
            if c == 1:
                cell.fill = SPRINT_FILL
        ws.row_dimensions[r].height = 90

    # Re-aplicar validacao de Status (coluna E) cobrindo as novas linhas
    last_row = start_row + len(NOVAS_SPRINTS) - 1
    dv = DataValidation(type="list",
                        formula1=f'"{",".join(STATUS_OPTIONS)}"',
                        allow_blank=True)
    dv.add(f"E{start_row}:E{last_row}")
    ws.add_data_validation(dv)


def append_backlog_tasks(ws):
    """Anexa as novas tarefas na aba Backlog."""
    start_row = ws.max_row + 1
    for i, t in enumerate(NOVAS_TAREFAS):
        sprint, tid, tipo, titulo, desc, criterios, deps, testes, prio = t
        r = start_row + i
        ws.cell(row=r, column=1, value=sprint)
        ws.cell(row=r, column=2, value=tid)
        ws.cell(row=r, column=3, value=tipo)
        ws.cell(row=r, column=4, value=prio)
        ws.cell(row=r, column=5, value=titulo)
        ws.cell(row=r, column=6, value=desc)
        ws.cell(row=r, column=7, value=criterios)
        ws.cell(row=r, column=8, value=deps)
        ws.cell(row=r, column=9, value=testes)
        ws.cell(row=r, column=10, value="Pendente")
        ws.cell(row=r, column=11, value="")
        ws.cell(row=r, column=12, value="")
        ws.cell(row=r, column=13, value="FASE 2 — análise PO + UX Designer + UI Designer (2026-05-24)")
        ws.cell(row=r, column=14, value="Não")
        ws.cell(row=r, column=15, value="")
        for c in range(1, 16):
            style_body(ws.cell(row=r, column=c), highlight=True)
        ws.row_dimensions[r].height = 130

    last_row = start_row + len(NOVAS_TAREFAS) - 1

    # Re-aplicar dropdowns nas novas linhas
    dv_status = DataValidation(type="list",
                               formula1=f'"{",".join(STATUS_OPTIONS)}"',
                               allow_blank=True)
    dv_status.add(f"J{start_row}:J{last_row}")
    ws.add_data_validation(dv_status)

    dv_prio = DataValidation(type="list",
                             formula1=f'"{",".join(PRIORIDADE_OPTIONS)}"',
                             allow_blank=True)
    dv_prio.add(f"D{start_row}:D{last_row}")
    ws.add_data_validation(dv_prio)

    dv_tipo = DataValidation(type="list",
                             formula1=f'"{",".join(TIPO_OPTIONS)}"',
                             allow_blank=True)
    dv_tipo.add(f"C{start_row}:C{last_row}")
    ws.add_data_validation(dv_tipo)

    dv_imp = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True)
    dv_imp.add(f"N{start_row}:N{last_row}")
    ws.add_data_validation(dv_imp)

    # Estende auto_filter ao novo intervalo
    ws.auto_filter.ref = f"A1:O{last_row}"


def append_test_plan(ws):
    """Anexa novos itens de teste na aba Plano de Testes."""
    start_row = ws.max_row + 1
    for i, t in enumerate(NOVOS_TESTES):
        sprint_alvo, modulo, tipo, desc = t
        r = start_row + i
        ws.cell(row=r, column=1, value=sprint_alvo)
        ws.cell(row=r, column=2, value=modulo)
        ws.cell(row=r, column=3, value=tipo)
        ws.cell(row=r, column=4, value=desc)
        ws.cell(row=r, column=5, value="")
        ws.cell(row=r, column=6, value="Pendente")
        ws.cell(row=r, column=7, value="")
        ws.cell(row=r, column=8, value="FASE 2 — adicionado em 2026-05-24 (análise PO/UX/UI)")
        for c in range(1, 9):
            style_body(ws.cell(row=r, column=c), highlight=True)
        ws.row_dimensions[r].height = 35

    last_row = start_row + len(NOVOS_TESTES) - 1
    dv = DataValidation(type="list",
                        formula1=f'"{",".join(STATUS_OPTIONS)}"',
                        allow_blank=True)
    dv.add(f"F{start_row}:F{last_row}")
    ws.add_data_validation(dv)

    ws.auto_filter.ref = f"A1:H{last_row}"


def add_decision_entry(ws):
    """Registra a decisao na aba Decisoes para deixar rastro."""
    start_row = ws.max_row + 1
    # Procura proximo ID
    next_id = "DEC-001"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] and isinstance(row[0], str) and row[0].startswith("DEC-"):
            try:
                num = int(row[0].split("-")[1]) + 1
                next_id = f"DEC-{num:03d}"
            except (IndexError, ValueError):
                pass

    contexto = ("ABERTURA DA FASE 2 DO PROJETO. A Fase 1 (Sprints 0 a 16) "
                "entregou o MVP com backend completo, mobile com a maior parte "
                "das telas e painel Streamlit com login/relatórios. Em 2026-05-24, "
                "análise conduzida pelos papéis Product Owner (PO), UX Designer "
                "e UI Designer identificou três problemas críticos: (P1) não "
                "existe tela de gestão de Contas em lugar nenhum — backend tem "
                "CRUD mas mobile e painel não expõem; (P2) painel é só leitura, "
                "sem cadastrar/editar/excluir transação; (P3) Dashboard mobile "
                "não tem atalho para a ação primária (Nova Transação). Outros "
                "problemas de menor severidade catalogados como P4 a P12 em "
                "analise_po_ux_ui_orcafacil.md.")
    decisao = ("Abrir FASE 2 do projeto com 3 sprints novas: Sprint 17 (Gestão "
               "de Contas — fecha P1), Sprint 18 (Painel completo — fecha P2 "
               "e P5) e Sprint 19 (Polimento UX/UI — fecha P3, P4, P6, P7, P8, "
               "P9). Total de 16 novas tarefas + 17 itens de teste. Marcar "
               "todas como Pendentes; preservar integralmente o tracking das "
               "sprints da Fase 1.")
    alternativas = ("Alt. 1: refazer a planilha do zero (descartada: destruiria "
                    "o tracking das 64 tarefas Concluídas da Fase 1). "
                    "Alt. 2: misturar essas tarefas dentro de sprints já "
                    "existentes (descartada: confundiria o chat executor e "
                    "diluiria o conceito de fase).")
    justificativa = ("Preservar a planilha como fonte única de verdade e dar "
                     "continuidade incremental. Sprints novas seguem o mesmo "
                     "padrão da Fase 1 (descrição + critérios + dependências + "
                     "testes obrigatórios). Nome das sprints prefixado com "
                     "[FASE 2] para deixar a separação visual evidente.")
    impacto = ("Chat executor pega S17-T01 como próxima tarefa Pendente. "
               "PROMPT_INICIAL_CLAUDE_CODE.md atualizado para refletir a Fase 2 "
               "e o histórico da Fase 1. Análise completa em "
               "analise_po_ux_ui_orcafacil.md. Linhas da Fase 2 destacadas "
               "em amarelo claro na planilha para facilitar identificação.")

    ws.cell(row=start_row, column=1, value=next_id)
    ws.cell(row=start_row, column=2, value="2026-05-24")
    ws.cell(row=start_row, column=3, value="17-19")
    ws.cell(row=start_row, column=4, value=contexto)
    ws.cell(row=start_row, column=5, value=decisao)
    ws.cell(row=start_row, column=6, value=alternativas)
    ws.cell(row=start_row, column=7, value=justificativa)
    ws.cell(row=start_row, column=8, value=impacto)
    for c in range(1, 9):
        style_body(ws.cell(row=start_row, column=c), highlight=True)
    ws.row_dimensions[start_row].height = 100


def append_fase2_banner_to_instructions(ws):
    """Anexa banner FASE 2 ao final da aba Instrucoes."""
    row = ws.max_row + 2
    title_font = Font(name="Calibri", size=14, bold=True, color="C00000")
    body_font = Font(name="Calibri", size=11, color="000000")
    note_font = Font(name="Calibri", size=10, italic=True, color="595959")

    lines = [
        ("FASE 2 — Aberta em 2026-05-24", title_font, 26),
        ("A Fase 1 (Sprints 0 a 16) entregou o MVP do OrçaFácil: backend "
         "completo, mobile com a maior parte das telas, painel Streamlit com "
         "login e relatórios. Análise pelos papéis Product Owner, UX Designer "
         "e UI Designer (documento analise_po_ux_ui_orcafacil.md) identificou "
         "lacunas e atrito que motivaram a abertura da Fase 2.", body_font, 80),
        ("Fase 2 = Sprints 17, 18 e 19 (linhas destacadas em amarelo claro "
         "nas abas Visão Geral, Backlog e Plano de Testes).", body_font, 35),
        ("Sprint 17 — Gestão de Contas (mobile + painel): fecha o bloqueio "
         "crítico de não haver tela para criar conta.", body_font, 35),
        ("Sprint 18 — Painel web com cadastro completo: paridade funcional "
         "entre painel e mobile (cadastrar/editar/excluir transação, importar "
         "CSV).", body_font, 35),
        ("Sprint 19 — Polimento UX/UI: atalho de Nova Transação no Dashboard, "
         "onboarding pós-cadastro, tokens de tema, dark mode real, item ativo "
         "no drawer, snackbar de saldo, picker de ícone.", body_font, 50),
        ("Regras inegociáveis seguem valendo: testes obrigatórios antes de "
         "concluir, regressão de saldo sempre verde, sem decisões fora da "
         "planilha. A bateria S05-T05 deve continuar verde após cada sprint "
         "da Fase 2.", body_font, 55),
        ("Justificativa registrada em DEC-001 na aba Decisões.",
         note_font, 22),
    ]

    for text, font, height in lines:
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = font
        cell.alignment = WRAP
        ws.row_dimensions[row].height = height
        row += 1


def main():
    wb = load_workbook(XLSX_PATH)
    append_sprints_overview(wb["Visão Geral"])
    append_backlog_tasks(wb["Backlog"])
    append_test_plan(wb["Plano de Testes"])
    add_decision_entry(wb["Decisões"])
    append_fase2_banner_to_instructions(wb["Instruções"])
    wb.save(XLSX_PATH)
    print(f"Planilha atualizada: {XLSX_PATH}")
    print(f"Novas sprints (Fase 2): {len(NOVAS_SPRINTS)}")
    print(f"Novas tarefas (Fase 2): {len(NOVAS_TAREFAS)}")
    print(f"Novos itens de teste (Fase 2): {len(NOVOS_TESTES)}")


if __name__ == "__main__":
    main()
