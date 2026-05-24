"""
Anexa Sprints 20-22 (Fase 3) + Sprint 23 (opcional / backlog) na planilha,
preservando tracking das Fases 1 e 2.

Tambem registra DEC-003 — politica de auto-commit/push de documentacao
ao fim de cada fase a partir da Fase 3.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


XLSX_PATH = r"d:\Projetos\OrcamentoFamiliar\orcafacil_planejamento_sprints.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
SPRINT_FILL = PatternFill("solid", fgColor="D9E1F2")
F3_HIGHLIGHT = PatternFill("solid", fgColor="E2EFDA")  # verde claro = Fase 3

STATUS_OPTIONS = [
    "Pendente", "Em andamento", "Bloqueada", "Em teste", "Concluída", "Cancelada",
]
PRIORIDADE_OPTIONS = ["Alta", "Média", "Baixa"]
TIPO_OPTIONS = ["Backend", "Frontend", "Mobile", "Infra", "Banco", "Teste",
                "Documentação", "Deploy"]


FASE_TAG = "[FASE 3]"
FASE_OBS = ("FASE 3 do projeto — Sprints 0-16 = FASE 1 (MVP); 17-19 = FASE 2 "
            "(UX/UI); 20+ = FASE 3 (hardening tecnico). Analise por papeis: "
            "Software Architect, QA Engineer, Security Engineer e DevOps "
            "Engineer (2026-05-24). Documento-base: "
            "analise_architect_qa_security_devops_orcafacil.md.")


NOVAS_SPRINTS = [
    {
        "numero": 20,
        "nome": f"{FASE_TAG} Sprint 20 — Hardening de Segurança e Dependências",
        "objetivo": "Fechar os achados criticos/altos da analise Security: sem "
                    "varredura de CVE em deps, rate limit so em /auth, ausencia "
                    "de headers de seguranca HTTP, refresh token sem revogacao, "
                    "politica de senha frouxa.",
        "entregaveis": "Dependabot ativo nos 2 repos; pip-audit no CI backend; "
                       "rate limit estendido por usuario nas mutacoes e em "
                       "/imports/csv; middleware de headers de seguranca; "
                       "revogacao de refresh token + endpoint /auth/logout; "
                       "politica de senha forte no schema.",
    },
    {
        "numero": 21,
        "nome": f"{FASE_TAG} Sprint 21 — Pipeline Mobile e Observabilidade",
        "objetivo": "Fechar o gap critico de CI no mobile e o gap alto de "
                    "observabilidade no backend para deploy futuro fazer sentido.",
        "entregaveis": "GitHub Actions no repo Flutter (analyze + test + build "
                       "smoke) com cobertura >=70%; logging agregado externo "
                       "(Better Stack ou equivalente); endpoint /healthz com "
                       "DB ping; endpoint /metrics Prometheus-compatible.",
    },
    {
        "numero": 22,
        "nome": f"{FASE_TAG} Sprint 22 — Deploy e Backup (retomar S16)",
        "objetivo": "Retomar as 3 tarefas canceladas da Sprint 16 (Supabase, "
                    "deploy, APK) somadas a backup automatizado. Ao final desta "
                    "sprint o produto esta efetivamente em producao publica.",
        "entregaveis": "Postgres gerenciado provisionado e com migrations "
                       "aplicadas; backend deployado (Render/Railway) com smoke "
                       "verde; painel Streamlit deployado; APK release assinado; "
                       "backup automatizado validado por restore manual; "
                       "docs/DEPLOY.md.",
    },
    {
        "numero": 23,
        "nome": f"{FASE_TAG} Sprint 23 — Robustez ponta a ponta (OPCIONAL)",
        "objetivo": "Sprint opcional/backlog. Aborda achados medio/baixos da "
                    "analise: ausencia de versionamento de API, contratos "
                    "Dart manuais, sem E2E cross-tier, sem fixtures de demo, "
                    "sem runbook.",
        "entregaveis": "Prefixo /v1/ em todos endpoints; cliente Dart gerado "
                       "do OpenAPI no CI mobile; suite E2E minima via Playwright; "
                       "scripts/seed_demo.py; docs/RUNBOOK.md.",
    },
]


# (sprint, id, tipo, titulo, descricao, criterios, deps, testes, prioridade)
NOVAS_TAREFAS = [
    # ---------- Sprint 20 — Hardening de Segurança ----------
    (20, "S20-T01", "Infra",
     "Habilitar Dependabot nos 2 repos",
     "Criar .github/dependabot.yml no backend_OrcamentoFacil e em "
     "frontend_OrcamentoFacil. Configurar grupos: prod, dev e actions. "
     "Frequencia weekly. PRs auto-criados para vulnerabilidades High/Critical.",
     "Dependabot habilitado nos 2 repos; primeira varredura roda em <24h; "
     "PRs aparecem quando ha vulneravel.",
     "—",
     "Validacao: rodar manualmente o primeiro check; criar PR de teste com "
     "dep desatualizada e validar deteccao.",
     "Alta"),
    (20, "S20-T02", "Infra",
     "pip-audit como step obrigatorio no CI backend",
     "Adicionar job no .github/workflows/ci.yml que roda pip-audit contra "
     "requirements.txt. Falha vermelha em vulnerabilidades High ou Critical; "
     "warning em Medium.",
     "CI quebra propositalmente em High/Critical e passa quando nao ha; "
     "documentado em README como remediar.",
     "S20-T01",
     "Teste manual: injetar dep com CVE conhecido e validar falha do CI.",
     "Alta"),
    (20, "S20-T03", "Backend",
     "Rate limiting estendido por usuario em mutacoes e /imports",
     "Configurar slowapi com chave por usuario autenticado (nao so IP). "
     "Limites: /imports/csv 1/min/user; /transactions POST/PATCH/DELETE "
     "60/min/user; /accounts POST/PATCH/DELETE 10/min/user; /budgets e /goals "
     "10/min/user.",
     "Excesso retorna 429 com Retry-After; ratelimit nao impacta GET; metricas "
     "do slowapi expostas (preparando S21-T05).",
     "—",
     "Integration tests: estourar limite em cada endpoint listado e validar "
     "429; GET nao e afetado; limite reseta apos janela.",
     "Alta"),
    (20, "S20-T04", "Backend",
     "Middleware de headers de seguranca HTTP",
     "Adicionar middleware FastAPI (proprio ou secure-headers) que injeta: "
     "Content-Security-Policy (default-src 'self'), Strict-Transport-Security "
     "(em prod), X-Content-Type-Options (nosniff), Referrer-Policy "
     "(strict-origin-when-cross-origin), X-Frame-Options (DENY).",
     "Response de qualquer endpoint inclui os 5 headers; testes validam "
     "presenca; HSTS so e enviado quando ENVIRONMENT=production.",
     "—",
     "Integration tests: GET /health valida headers; teste especifico para "
     "HSTS so em prod.",
     "Alta"),
    (20, "S20-T05", "Backend",
     "Revogacao de refresh token + endpoint /auth/logout",
     "Migration: criar tabela revoked_tokens(jti, user_id, revoked_at). "
     "Endpoint POST /auth/logout (autenticado) que adiciona o jti do refresh "
     "atual. Em /auth/refresh, checar revoked_tokens antes de emitir novo "
     "access. Cleanup periodico (cron ou on-demand) de tokens expirados.",
     "Logout revoga corretamente; refresh com token revogado retorna 401; "
     "tokens expirados nao acumulam para sempre.",
     "—",
     "Integration tests: fluxo register -> login -> refresh OK -> logout -> "
     "refresh 401; teste de cleanup; teste de logout sem refresh ativo.",
     "Média"),
    (20, "S20-T06", "Backend",
     "Politica de senha forte no UserCreate",
     "Validator no schema Pydantic UserCreate: minimo 8 chars, ao menos 1 "
     "letra e 1 numero. Mensagem de erro clara em portugues.",
     "Senha fraca retorna 422 com mensagem util; senha forte cria usuario.",
     "—",
     "Unit tests dos schemas cobrindo casos: vazia, curta, so letras, so "
     "numeros, forte.",
     "Média"),

    # ---------- Sprint 21 — Pipeline Mobile + Observabilidade ----------
    (21, "S21-T01", "Mobile",
     "GitHub Actions no repo frontend_OrcamentoFacil",
     "Criar .github/workflows/ci.yml: setup-flutter@v2, flutter pub get, "
     "flutter analyze (gate), flutter test --coverage. Gate de cobertura "
     ">=70% via lcov-summary ou similar.",
     "Workflow verde em push de branch nova; falha vermelha em analyze "
     "warning ou cobertura abaixo de 70%; badge no README.",
     "—",
     "Validar com PR de teste quebrando analyze e PR de teste reduzindo "
     "cobertura.",
     "Alta"),
    (21, "S21-T02", "Mobile",
     "Job CI: flutter build apk --debug",
     "Adicionar job paralelo ao ci.yml que roda flutter build apk --debug "
     "para garantir que o app compila. Artefato APK guardado por 7 dias.",
     "Build smoke verde a cada push; APK debug baixavel da Actions UI.",
     "S21-T01",
     "Validacao manual: baixar APK de um run, instalar em emulador, abrir.",
     "Média"),
    (21, "S21-T03", "Infra",
     "Logging agregado externo",
     "Configurar Better Stack (ou Logtail/Grafana Cloud — free tier suficiente). "
     "Backend envia logs estruturados via handler dedicado quando "
     "LOG_SHIPPING_URL e LOG_SHIPPING_TOKEN estao definidos. Local development "
     "continua so com stdout.",
     "Logs do backend aparecem no painel externo com timestamp, level, "
     "request_id e payload. Em dev (sem vars setadas), comportamento atual "
     "preservado.",
     "—",
     "Integration test do handler: com vars setadas, log e enviado; sem vars, "
     "fica so em stdout.",
     "Alta"),
    (21, "S21-T04", "Backend",
     "Endpoint /healthz com DB ping",
     "POST /healthz retorna {status, version, uptime_seconds, "
     "database: {status, latency_ms}}. Distinto de /health (que so retorna "
     "{status: ok}). Usa SELECT 1 + timer.",
     "DB OK retorna 200 com latency; DB down retorna 503 com detalhe; nao "
     "expoe senha nem strings de conexao.",
     "—",
     "Integration tests: DB OK; DB down (mock); response time razoavel.",
     "Média"),
    (21, "S21-T05", "Backend",
     "Endpoint /metrics Prometheus-compatible",
     "Usar prometheus-fastapi-instrumentator. Expoe latencia por endpoint, "
     "request count, error count. Endpoint /metrics sem autenticacao (default) "
     "ou behind basic auth via env.",
     "GET /metrics retorna formato texto Prometheus parseavel; metricas "
     "aumentam com requests.",
     "—",
     "Integration tests: fazer 3 requests; checar que counter incrementou.",
     "Média"),

    # ---------- Sprint 22 — Deploy e Backup ----------
    (22, "S22-T01", "Deploy",
     "Provisionar Postgres gerenciado (Supabase ou Neon)",
     "Criar projeto Supabase (free tier) OU Neon Postgres. Anotar "
     "DATABASE_URL (postgresql+psycopg://...). Aplicar todas as migrations "
     "via alembic upgrade head remotamente. Atualizar .env.example "
     "indicando o provedor escolhido.",
     "Migrations aplicadas com sucesso; SELECT 1 remoto responde; tabelas "
     "existem.",
     "—",
     "Smoke remoto: criar user via /auth/register contra DATABASE_URL "
     "gerenciado.",
     "Alta"),
    (22, "S22-T02", "Deploy",
     "Deploy backend em Render ou Railway",
     "Escolher Render (recomendado) ou Railway. Configurar build a partir "
     "do main do GitHub. Secrets: DATABASE_URL, JWT_SECRET, CORS_ORIGINS, "
     "LOG_SHIPPING_*. Healthcheck /healthz. Auto-deploy em merge para main.",
     "Backend responde 200 em /healthz publico via HTTPS; auto-deploy "
     "funciona em commit dummy.",
     "S22-T01",
     "scripts/smoke_prod.py contra URL publica: register + login + me + "
     "criar conta + criar transacao + listar.",
     "Alta"),
    (22, "S22-T03", "Deploy",
     "Deploy do painel Streamlit",
     "Streamlit Community Cloud (free) ou Render. ORCAFACIL_API_URL aponta "
     "para o backend de prod (S22-T02). Login funcional.",
     "Painel acessivel via URL publica HTTPS; login contra prod funciona; "
     "todas as paginas carregam.",
     "S22-T02",
     "Smoke manual: login, criar conta, criar transacao, ver no relatorio.",
     "Alta"),
    (22, "S22-T04", "Mobile",
     "APK release apontando para producao",
     "Configurar build com --dart-define=API_BASE_URL=<prod>. Gerar keystore "
     "(release.keystore) e build.gradle assinando. flutter build apk --release. "
     "Documentar processo em frontend/README.md (sem expor keystore senha).",
     "APK release gera sem erro; instala em device/emulador Android; login "
     "contra backend de producao funciona.",
     "S22-T02",
     "Validacao manual em emulador: fluxo cadastro -> login -> criar conta -> "
     "criar transacao -> dashboard.",
     "Média"),
    (22, "S22-T05", "Infra",
     "Backup automatizado do Postgres",
     "Usar backup nativo do Supabase (point-in-time recovery free tier) OU "
     "GitHub Actions cron diario que faz pg_dump e envia para B2/S3 "
     "(criptografado). Retencao minima 7 dias.",
     "Backup roda diariamente sem intervencao; arquivo restaurado em DB de "
     "teste valida integridade.",
     "S22-T01",
     "Restore manual em DB de teste local: aplicar dump; SELECT count(*) "
     "bate com producao.",
     "Alta"),
    (22, "S22-T06", "Documentação",
     "docs/DEPLOY.md",
     "Documentar: arquitetura de prod, provedores escolhidos, como acessar "
     "logs e metricas, como rotacionar JWT_SECRET, como fazer rollback, "
     "como restaurar backup. NAO incluir secrets no doc — apenas onde "
     "encontra-los.",
     "Documento completo; novo dev consegue entender e operar a partir dele.",
     "S22-T05",
     "Revisao: seguir o doc do zero (em outra maquina/sessao) para validar.",
     "Média"),

    # ---------- Sprint 23 — Robustez ponta a ponta (OPCIONAL) ----------
    (23, "S23-T01", "Backend",
     "Prefixo /v1/ em todos os endpoints",
     "Migrar todos os routers para APIRouter(prefix='/v1'). Atualizar mobile "
     "(api_client.dart base) e painel (api.py base). Atualizar smoke_prod.py. "
     "Manter compatibilidade temporaria via redirect /auth -> /v1/auth durante "
     "1 release.",
     "Todos os endpoints servem em /v1/; redirects funcionam; clientes "
     "atualizados; openapi.json reflete prefixo.",
     "—",
     "Integration tests adaptados; smoke verde antes e depois do redirect ser "
     "removido.",
     "Média"),
    (23, "S23-T02", "Mobile",
     "Gerar cliente Dart do OpenAPI no CI",
     "Adicionar job no CI mobile: baixar openapi.json do staging, rodar "
     "openapi-generator-cli gerar cliente Dart, comparar com a versao "
     "checkada. Diff = fail. Substituir clients manuais por gerados "
     "progressivamente.",
     "CI quebra quando cliente checkado diverge do OpenAPI atual.",
     "S23-T01",
     "PR de teste: mudar campo no Pydantic, validar CI vermelho ate "
     "regenerar cliente.",
     "Média"),
    (23, "S23-T03", "Teste",
     "Suite E2E minima cross-tier (Playwright)",
     "Configurar Playwright (Python) rodando contra painel deployado em "
     "staging. Cenarios: login, criar conta, criar transacao, ver no "
     "relatorio. Job no CI dispara apos deploy de staging.",
     "Suite verde em staging; quebra real quando algo da pipeline falha.",
     "S22-T03",
     "Adicionar regressao proposital (mudar texto) para validar deteccao.",
     "Média"),
    (23, "S23-T04", "Backend",
     "scripts/seed_demo.py",
     "Script que cria 1 user demo (com senha previsivel para demo), 3 contas, "
     "8 categorias, 60 transacoes distribuidas em 90 dias com mix realista de "
     "receitas e despesas. Idempotente: roda 2 vezes sem duplicar.",
     "Rodar o script em DB limpo produz cenario navegavel imediatamente; "
     "rodar denovo nao duplica.",
     "—",
     "Unit test do script (mock DB); smoke manual contra DB local.",
     "Baixa"),
    (23, "S23-T05", "Documentação",
     "docs/RUNBOOK.md com 5 cenarios de incidente",
     "Documentar como agir em: (1) backend nao responde; (2) latencia "
     "anormal; (3) disco do Postgres cheio; (4) migration falhou em prod; "
     "(5) suspeita de comprometimento de credenciais.",
     "Cada cenario tem: sintomas, comandos de diagnostico, passos de "
     "remediacao, links uteis.",
     "S22-T06",
     "Revisao: simular mentalmente cada cenario seguindo o runbook.",
     "Baixa"),
]


# (sprint_alvo, modulo, tipo_teste, descricao)
NOVOS_TESTES = [
    (20, "deps", "Manual", "Dependabot dispara PR para vuln injetada propositalmente"),
    (20, "ci", "Integração", "pip-audit quebra CI em vuln High/Critical"),
    (20, "rate-limit", "Integração", "/imports/csv 429 ao estourar 1/min; reset apos janela"),
    (20, "rate-limit", "Integração", "/transactions POST 429 ao estourar 60/min por user"),
    (20, "headers", "Integração", "5 security headers presentes em todas as responses"),
    (20, "headers", "Integração", "HSTS so em ENVIRONMENT=production"),
    (20, "auth", "Integração", "Refresh com token revogado retorna 401"),
    (20, "auth", "Integração", "Logout revoga refresh atual"),
    (20, "auth", "Unitário", "Politica de senha rejeita fraca / aceita forte"),

    (21, "ci-mobile", "Manual", "Workflow Flutter verde em push; quebra em analyze warning"),
    (21, "ci-mobile", "Manual", "Gate de cobertura 70% quebra com PR reduzindo cobertura"),
    (21, "ci-mobile", "Manual", "Build APK debug verde; artefato disponivel"),
    (21, "logging", "Integração", "Logs enviados ao agregador quando env vars setadas"),
    (21, "logging", "Integração", "Sem env vars, comportamento atual preservado (stdout)"),
    (21, "/healthz", "Integração", "/healthz retorna status DB e latencia"),
    (21, "/healthz", "Integração", "/healthz retorna 503 quando DB down"),
    (21, "/metrics", "Integração", "/metrics expoe counters Prometheus apos requests"),

    (22, "deploy-db", "Smoke", "alembic upgrade head contra DB gerenciado"),
    (22, "deploy-db", "Smoke", "register + login contra DATABASE_URL remoto"),
    (22, "deploy-api", "Smoke", "smoke_prod.py contra URL publica verde"),
    (22, "deploy-api", "Manual", "Auto-deploy dispara em commit para main"),
    (22, "deploy-panel", "Manual", "Painel publico carrega e loga contra prod"),
    (22, "deploy-mobile", "Manual", "APK release instala e loga contra prod"),
    (22, "backup", "Manual", "Restore do ultimo backup em DB de teste local"),

    (23, "/v1", "Integração", "Todos endpoints servem em /v1; redirects funcionam"),
    (23, "client-dart", "Manual", "Cliente Dart gerado diff-clean apos regeneracao"),
    (23, "e2e", "E2E", "Playwright: login -> criar conta -> criar transacao -> ver relatorio"),
    (23, "seed", "Unitário", "seed_demo idempotente (rodar 2x nao duplica)"),
    (23, "runbook", "Manual", "5 cenarios revistos mentalmente e validados"),
]


# =========================================================
# Helpers
# =========================================================

def style_body(cell, highlight=False):
    cell.alignment = WRAP
    cell.border = BORDER
    if highlight:
        cell.fill = F3_HIGHLIGHT


def append_sprints_overview(ws):
    start_row = ws.max_row + 1
    for i, s in enumerate(NOVAS_SPRINTS):
        r = start_row + i
        ws.cell(row=r, column=1, value=s["numero"])
        ws.cell(row=r, column=2, value=s["nome"])
        ws.cell(row=r, column=3, value=s["objetivo"])
        ws.cell(row=r, column=4, value=s["entregaveis"])
        ws.cell(row=r, column=5, value="Pendente")
        ws.cell(row=r, column=6, value="")
        ws.cell(row=r, column=7, value="")
        ws.cell(row=r, column=8, value=FASE_OBS)
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            style_body(cell, highlight=True)
            if c == 1:
                cell.fill = SPRINT_FILL
        ws.row_dimensions[r].height = 95
    last_row = start_row + len(NOVAS_SPRINTS) - 1
    dv = DataValidation(type="list",
                        formula1=f'"{",".join(STATUS_OPTIONS)}"',
                        allow_blank=True)
    dv.add(f"E{start_row}:E{last_row}")
    ws.add_data_validation(dv)


def append_backlog_tasks(ws):
    start_row = ws.max_row + 1
    for i, t in enumerate(NOVAS_TAREFAS):
        sprint, tid, tipo, titulo, desc, criterios, deps, testes, prio = t
        r = start_row + i
        ws.cell(row=r, column=1, value=sprint)
        ws.cell(row=r, column=2, value=tid)
        ws.cell(row=r, column=3, value=tipo)
        ws.cell(row=r, column=4, value=prio)
        ws.cell(row=r, column=5, value=titulo)
        ws.cell(row=r, column=6, value=desc)
        ws.cell(row=r, column=7, value=criterios)
        ws.cell(row=r, column=8, value=deps)
        ws.cell(row=r, column=9, value=testes)
        ws.cell(row=r, column=10, value="Pendente")
        ws.cell(row=r, column=11, value="")
        ws.cell(row=r, column=12, value="")
        ws.cell(row=r, column=13,
                value="FASE 3 — analise Architect + QA + Security + DevOps (2026-05-24)")
        ws.cell(row=r, column=14, value="Não")
        ws.cell(row=r, column=15, value="")
        for c in range(1, 16):
            style_body(ws.cell(row=r, column=c), highlight=True)
        ws.row_dimensions[r].height = 130
    last_row = start_row + len(NOVAS_TAREFAS) - 1
    dv_s = DataValidation(type="list",
                          formula1=f'"{",".join(STATUS_OPTIONS)}"',
                          allow_blank=True)
    dv_s.add(f"J{start_row}:J{last_row}")
    ws.add_data_validation(dv_s)
    dv_p = DataValidation(type="list",
                          formula1=f'"{",".join(PRIORIDADE_OPTIONS)}"',
                          allow_blank=True)
    dv_p.add(f"D{start_row}:D{last_row}")
    ws.add_data_validation(dv_p)
    dv_t = DataValidation(type="list",
                          formula1=f'"{",".join(TIPO_OPTIONS)}"',
                          allow_blank=True)
    dv_t.add(f"C{start_row}:C{last_row}")
    ws.add_data_validation(dv_t)
    dv_i = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True)
    dv_i.add(f"N{start_row}:N{last_row}")
    ws.add_data_validation(dv_i)
    ws.auto_filter.ref = f"A1:O{last_row}"


def append_test_plan(ws):
    start_row = ws.max_row + 1
    for i, t in enumerate(NOVOS_TESTES):
        sprint_alvo, modulo, tipo, desc = t
        r = start_row + i
        ws.cell(row=r, column=1, value=sprint_alvo)
        ws.cell(row=r, column=2, value=modulo)
        ws.cell(row=r, column=3, value=tipo)
        ws.cell(row=r, column=4, value=desc)
        ws.cell(row=r, column=5, value="")
        ws.cell(row=r, column=6, value="Pendente")
        ws.cell(row=r, column=7, value="")
        ws.cell(row=r, column=8,
                value="FASE 3 — adicionado em 2026-05-24")
        for c in range(1, 9):
            style_body(ws.cell(row=r, column=c), highlight=True)
        ws.row_dimensions[r].height = 35
    last_row = start_row + len(NOVOS_TESTES) - 1
    dv = DataValidation(type="list",
                        formula1=f'"{",".join(STATUS_OPTIONS)}"',
                        allow_blank=True)
    dv.add(f"F{start_row}:F{last_row}")
    ws.add_data_validation(dv)
    ws.auto_filter.ref = f"A1:H{last_row}"


def add_decision_entry(ws):
    """Registra DEC-003 — politica de auto-commit/push de docs."""
    start_row = ws.max_row + 1

    contexto = ("Apos a Fase 2, formalizamos o ciclo: Fase concluida -> "
                "skill portfolio-docs re-renderiza a planilha em docs/sprints_*.md "
                "-> commit no repo de metodologia -> push para o GitHub. Ate "
                "agora isso foi feito manualmente pelo humano. A partir da Fase "
                "3, queremos que o proprio chat executor faca essa rotina ao "
                "final de cada fase, mantendo a 'documentacao viva' realmente "
                "viva sem depender de intervencao manual.")
    decisao = ("Ao concluir TODAS as tarefas de uma fase, o agente executor "
               "deve: (1) confirmar que a planilha esta atualizada (todas as "
               "tarefas Concluida ou Cancelada); (2) rodar o script render que "
               "regera os docs MD; (3) atualizar README.md com o novo status; "
               "(4) git add + git commit com mensagem padronizada "
               "'docs: fechamento da Fase N'; (5) PARAR e PEDIR ao humano "
               "confirmacao explicita antes de fazer git push.")
    alternativas = ("Alt. 1: push automatico sem confirmacao (descartada: perde "
                    "a barreira final, risco de empurrar coisa nao-saneada). "
                    "Alt. 2: agente so prepara o commit, humano executa "
                    "(descartada: anula a automacao). Alt. 3: manter manual "
                    "como na Fase 2 (descartada: nao escala e contradiz o "
                    "principio de documentacao viva).")
    justificativa = ("Confirmacao humana antes do push e a barreira certa: "
                     "automacao do trabalho repetitivo (render, add, commit) "
                     "+ humano no controle do que vai a publico. Custo do "
                     "humano e 5 segundos; valor e zero risco de publicar "
                     "credencial/dado sensivel acidentalmente.")
    impacto = ("PROMPT_INICIAL_CLAUDE_CODE.md da Fase 3 inclui essa rotina "
               "como obrigatoria no protocolo de fim-de-fase. Agente da Fase "
               "3 deve verificar essa rotina antes de declarar a fase fechada. "
               "Esta DEC-003 referenciada explicitamente no PROMPT.")

    ws.cell(row=start_row, column=1, value="DEC-003")
    ws.cell(row=start_row, column=2, value="2026-05-24")
    ws.cell(row=start_row, column=3, value="20-23")
    ws.cell(row=start_row, column=4, value=contexto)
    ws.cell(row=start_row, column=5, value=decisao)
    ws.cell(row=start_row, column=6, value=alternativas)
    ws.cell(row=start_row, column=7, value=justificativa)
    ws.cell(row=start_row, column=8, value=impacto)
    for c in range(1, 9):
        style_body(ws.cell(row=start_row, column=c), highlight=True)
    ws.row_dimensions[start_row].height = 110


def append_fase3_banner_to_instructions(ws):
    row = ws.max_row + 2
    title_font = Font(name="Calibri", size=14, bold=True, color="385723")
    body_font = Font(name="Calibri", size=11, color="000000")
    note_font = Font(name="Calibri", size=10, italic=True, color="595959")

    lines = [
        ("FASE 3 — Aberta em 2026-05-24", title_font, 26),
        ("Apos Fase 1 (MVP, S0-16) e Fase 2 (UX/UI, S17-19), abrimos a Fase 3 "
         "para hardening tecnico identificado pelos papeis Software Architect, "
         "QA Engineer, Security Engineer e DevOps Engineer (documento "
         "analise_architect_qa_security_devops_orcafacil.md).", body_font, 60),
        ("Fase 3 = Sprints 20, 21, 22 (S23 opcional). Linhas destacadas em "
         "verde claro nas abas Visao Geral, Backlog e Plano de Testes.",
         body_font, 35),
        ("Sprint 20 — Hardening Seguranca e Dependencias: Dependabot, "
         "pip-audit, rate limit por user, headers de seguranca, refresh token "
         "revogavel, senha forte.", body_font, 50),
        ("Sprint 21 — Pipeline Mobile + Observabilidade: GitHub Actions "
         "Flutter, logging agregado, /healthz, /metrics.", body_font, 45),
        ("Sprint 22 — Deploy e Backup (retomar S16): Postgres gerenciado, "
         "backend e painel deployados, APK release, backup automatizado, "
         "docs/DEPLOY.md.", body_font, 50),
        ("Sprint 23 (OPCIONAL) — Robustez ponta a ponta: /v1, cliente Dart "
         "gerado, E2E Playwright, seed_demo, RUNBOOK.", body_font, 50),
        ("NOVA ROTINA OBRIGATORIA DE FIM-DE-FASE (DEC-003): ao concluir TODAS "
         "as tarefas de uma fase, o agente DEVE: (1) verificar planilha "
         "consistente; (2) re-renderizar docs/; (3) atualizar README.md com "
         "status; (4) git add + commit padronizado; (5) PARAR e pedir OK "
         "humano antes do push. Detalhes em DEC-003 (aba Decisoes) e no "
         "PROMPT_INICIAL.", body_font, 100),
        ("Regras inegociaveis seguem valendo: testes obrigatorios, regressao "
         "de saldo verde, sem decisoes fora da planilha. test_balance_"
         "regression_full_flow deve continuar verde apos cada sprint da "
         "Fase 3.", body_font, 55),
        ("Justificativa em DEC-003.", note_font, 22),
    ]

    for text, font, height in lines:
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = font
        cell.alignment = WRAP
        ws.row_dimensions[row].height = height
        row += 1


def main():
    wb = load_workbook(XLSX_PATH)
    append_sprints_overview(wb["Visão Geral"])
    append_backlog_tasks(wb["Backlog"])
    append_test_plan(wb["Plano de Testes"])
    add_decision_entry(wb["Decisões"])
    append_fase3_banner_to_instructions(wb["Instruções"])
    wb.save(XLSX_PATH)
    print(f"Planilha atualizada: {XLSX_PATH}")
    print(f"Novas sprints (Fase 3): {len(NOVAS_SPRINTS)} (S20-S22 + S23 opcional)")
    print(f"Novas tarefas (Fase 3): {len(NOVAS_TAREFAS)}")
    print(f"Novos itens de teste (Fase 3): {len(NOVOS_TESTES)}")


if __name__ == "__main__":
    main()
