"""
Gera a planilha XLSX de planejamento das sprints do projeto OrçaFácil.

A planilha é a "documentação verdade" que o Claude Code usará para executar
as tarefas em sequência, marcar progresso e registrar impedimentos.

Abas:
  1. Instruções        -> Como o Claude Code deve usar a planilha
  2. Visão Geral       -> Status macro de cada sprint
  3. Backlog           -> Backlog detalhado (uma linha por tarefa)
  4. Plano de Testes   -> Checklist de testes por feature
  5. Impedimentos      -> Log de bloqueios / dúvidas para o humano
  6. Decisões          -> Log de decisões técnicas tomadas
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


# ============================================================
# Estilos reutilizáveis
# ============================================================

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F4E78")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E78")
NOTE_FONT = Font(name="Calibri", size=10, italic=True, color="595959")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SPRINT_FILL = PatternFill("solid", fgColor="D9E1F2")

STATUS_OPTIONS = [
    "Pendente",
    "Em andamento",
    "Bloqueada",
    "Em teste",
    "Concluída",
    "Cancelada",
]

PRIORIDADE_OPTIONS = ["Alta", "Média", "Baixa"]
TIPO_OPTIONS = ["Backend", "Frontend", "Mobile", "Infra", "Banco", "Teste", "Documentação", "Deploy"]


# ============================================================
# Conteúdo: Sprints e Tarefas
# ============================================================

SPRINTS = [
    {
        "numero": 0,
        "nome": "Sprint 0 — Setup do Projeto e Infraestrutura",
        "objetivo": "Preparar o ambiente de desenvolvimento, repositório, Docker, "
                    "PostgreSQL local, configuração de ferramentas e CI mínimo. "
                    "Sem essa base nada do resto roda.",
        "entregaveis": "Repositório versionado, Docker Compose subindo Postgres, "
                       "estrutura de pastas backend/mobile/panel, pytest configurado, "
                       "pipeline CI rodando lint + testes em cada push.",
    },
    {
        "numero": 1,
        "nome": "Sprint 1 — Backend Core (FastAPI + DB)",
        "objetivo": "Subir o backend FastAPI com conexão ao Postgres via SQLAlchemy, "
                    "Alembic configurado, settings via Pydantic, logging e tratamento "
                    "global de erros.",
        "entregaveis": "API rodando localmente em /health, migrations criando o schema "
                       "inicial, logs estruturados, configurações via variáveis de ambiente.",
    },
    {
        "numero": 2,
        "nome": "Sprint 2 — Autenticação (JWT)",
        "objetivo": "Implementar cadastro, login, hash de senha, emissão e verificação "
                    "de JWT, dependency de usuário autenticado.",
        "entregaveis": "Endpoints /auth/register, /auth/login, /users/me funcionando "
                       "com testes unitários e de integração cobrindo cenários felizes "
                       "e de falha.",
    },
    {
        "numero": 3,
        "nome": "Sprint 3 — CRUD de Contas (Accounts)",
        "objetivo": "Gerenciar contas financeiras do usuário (conta corrente, poupança, "
                    "cartão, dinheiro). Saldo inicial e saldo atual.",
        "entregaveis": "CRUD completo de contas restrito ao usuário autenticado, "
                       "testes garantindo isolamento entre usuários.",
    },
    {
        "numero": 4,
        "nome": "Sprint 4 — CRUD de Categorias",
        "objetivo": "Gerenciar categorias de receita e despesa, com cor e ícone. "
                    "Semear categorias padrão para cada novo usuário.",
        "entregaveis": "CRUD de categorias, seed automático ao registrar usuário, "
                       "testes cobrindo seed e isolamento por usuário.",
    },
    {
        "numero": 5,
        "nome": "Sprint 5 — CRUD de Transações",
        "objetivo": "Registrar receitas e despesas, atualizando o saldo da conta. "
                    "Filtros, busca, paginação e ordenação.",
        "entregaveis": "CRUD completo de transações com atualização correta de saldo "
                       "(regra de negócio crítica), filtros e testes que impedem "
                       "regressão no cálculo de saldo.",
    },
    {
        "numero": 6,
        "nome": "Sprint 6 — Orçamentos Mensais",
        "objetivo": "Permitir definir limite mensal por categoria, calcular percentual "
                    "utilizado e disparar alertas (>80% e >100%).",
        "entregaveis": "Endpoints de orçamento, cálculo correto do percentual usado "
                       "no mês corrente, testes cobrindo alertas e mudança de mês.",
    },
    {
        "numero": 7,
        "nome": "Sprint 7 — Metas Financeiras",
        "objetivo": "Permitir criar metas (reserva, viagem, etc.) e atualizar progresso. "
                    "Marcar como concluída automaticamente.",
        "entregaveis": "CRUD de metas, atualização de progresso, regra automática de "
                       "conclusão (current_amount >= target_amount), testes cobrindo "
                       "transição de status.",
    },
    {
        "numero": 8,
        "nome": "Sprint 8 — Dashboard Financeiro (Backend)",
        "objetivo": "Endpoints de agregação: resumo mensal, breakdown por categoria, "
                    "fluxo de caixa.",
        "entregaveis": "Três endpoints de dashboard com queries otimizadas, testes "
                       "cobrindo períodos vazios, múltiplas contas e múltiplas categorias.",
    },
    {
        "numero": 9,
        "nome": "Sprint 9 — Importação CSV",
        "objetivo": "Permitir upload de extratos CSV com parsing e categorização "
                    "automática por palavras-chave.",
        "entregaveis": "Endpoint POST /imports/csv, parser robusto, regras de "
                       "categorização automática (UBER, IFOOD, NETFLIX, etc.), "
                       "criação em batch com atualização de saldos.",
    },
    {
        "numero": 10,
        "nome": "Sprint 10 — Hardening do Backend",
        "objetivo": "CORS, rate limiting, logs estruturados, OpenAPI completo, "
                    "cobertura de testes mínima 80%, refresh token.",
        "entregaveis": "Backend pronto para produção em termos de segurança e "
                       "observabilidade, com documentação OpenAPI navegável.",
    },
    {
        "numero": 11,
        "nome": "Sprint 11 — Flutter Setup e Autenticação",
        "objetivo": "Inicializar o app Flutter com Riverpod, go_router, Dio com "
                    "interceptor JWT e armazenamento seguro do token.",
        "entregaveis": "Telas Splash, Login e Cadastro funcionando contra o backend, "
                       "widget tests cobrindo formulários e navegação.",
    },
    {
        "numero": 12,
        "nome": "Sprint 12 — Flutter Dashboard e Transações",
        "objetivo": "Dashboard com cards e gráficos (fl_chart), lista de transações "
                    "com filtros, telas de criação e edição.",
        "entregaveis": "Fluxo completo de transações no app, widget tests com mocks "
                       "do Dio, integração visual com o backend.",
    },
    {
        "numero": 13,
        "nome": "Sprint 13 — Flutter Categorias, Orçamento e Metas",
        "objetivo": "Telas de categorias, orçamento (com alertas visuais) e metas "
                    "(com barra de progresso).",
        "entregaveis": "Telas funcionais consumindo os endpoints correspondentes, "
                       "widget tests cobrindo estados (vazio, com dados, erro).",
    },
    {
        "numero": 14,
        "nome": "Sprint 14 — Flutter Perfil, Configurações e Import CSV",
        "objetivo": "Telas de perfil, configurações e upload de CSV (file_picker).",
        "entregaveis": "App com todas as telas do MVP, fluxo de import CSV "
                       "funcionando ponta a ponta, testes cobrindo upload.",
    },
    {
        "numero": 15,
        "nome": "Sprint 15 — Painel Web (Streamlit)",
        "objetivo": "Painel web para visualização de relatórios mais ricos, "
                    "autenticado contra o mesmo backend.",
        "entregaveis": "Painel Streamlit com login, relatórios mensais/anuais, "
                       "gráficos analíticos e exportação básica.",
    },
    {
        "numero": 16,
        "nome": "Sprint 16 — Deploy e Documentação Final",
        "objetivo": "Deploy do backend, banco gerenciado, build do APK Android e "
                    "documentação consolidada do projeto.",
        "entregaveis": "Backend rodando em ambiente público (Railway/Render), "
                       "Supabase configurado, APK gerado, README e docs finais "
                       "concluídos, smoke tests em produção verdes.",
    },
]


# Cada item: (sprint, id_tarefa, tipo, titulo, descricao, criterios_aceitacao,
#             dependencias, testes_obrigatorios, prioridade)

TAREFAS = [
    # ---------------- Sprint 0 ----------------
    (0, "S00-T01", "Infra",
     "Criar repositório Git e estrutura inicial",
     "Inicializar repositório Git, criar pastas backend/, mobile/, panel/, docs/, scripts/. "
     "Adicionar .gitignore para Python, Flutter, IDEs e arquivos sensíveis. "
     "Criar README.md raiz descrevendo o projeto e como rodar.",
     "Repositório local criado, .gitignore cobrindo .env, __pycache__, build/, .dart_tool/, "
     ".idea/, .vscode/, *.log. README.md presente.",
     "—",
     "Validação manual: clonar em outra pasta e confirmar que arquivos sensíveis não entram.",
     "Alta"),
    (0, "S00-T02", "Infra",
     "Docker Compose com PostgreSQL local",
     "Criar docker-compose.yml com serviço postgres (versão 16), volume persistente, "
     "variáveis de ambiente via .env. Criar .env.example com placeholders.",
     "docker compose up sobe o Postgres na porta configurada, dados persistem entre restarts, "
     ".env.example documenta todas as variáveis.",
     "S00-T01",
     "Teste manual: subir o compose, conectar com psql/DBeaver, derrubar e subir de novo "
     "confirmando persistência.",
     "Alta"),
    (0, "S00-T03", "Backend",
     "Configurar ambiente Python e dependências",
     "Criar backend/pyproject.toml ou requirements.txt com FastAPI, SQLAlchemy, Alembic, "
     "Pydantic, Pydantic-Settings, Uvicorn, passlib[bcrypt], python-jose, pytest, pytest-asyncio, "
     "httpx, ruff, black. Configurar venv.",
     "pip install rodando sem erros, ruff e black configurados no pyproject.toml.",
     "S00-T01",
     "pip install -r requirements.txt sem erros; ruff check . sem violações no esqueleto.",
     "Alta"),
    (0, "S00-T04", "Teste",
     "Configurar pytest e coverage",
     "Criar backend/tests/, conftest.py, pytest.ini com configurações (asyncio_mode=auto, "
     "coverage mínimo 80%). Adicionar teste smoke que apenas valida import.",
     "pytest roda e exibe 1 teste passando; pytest --cov reporta cobertura.",
     "S00-T03",
     "Teste smoke: assert True; pytest --cov backend mostra relatório.",
     "Alta"),
    (0, "S00-T05", "Infra",
     "CI mínimo (GitHub Actions)",
     "Criar .github/workflows/ci.yml que roda ruff + pytest em cada push/PR. "
     "Service container do Postgres para integração.",
     "Workflow verde em push de branch nova; falha vermelha quando teste é quebrado de propósito.",
     "S00-T04",
     "Push de teste quebrando assert para verificar falha do CI; depois corrigir.",
     "Média"),
    (0, "S00-T06", "Documentação",
     "Documentar como rodar o projeto",
     "Atualizar README com seções: requisitos, setup, rodar backend, rodar testes, "
     "rodar Docker Compose, troubleshooting comum.",
     "Novo dev consegue subir o ambiente lendo só o README.",
     "S00-T02, S00-T03, S00-T04",
     "Revisão: seguir o README do zero em pasta nova.",
     "Média"),

    # ---------------- Sprint 1 ----------------
    (1, "S01-T01", "Backend",
     "Estrutura base do FastAPI",
     "Criar backend/app/main.py com FastAPI(), CORS middleware, exception handler global, "
     "endpoint /health retornando {status: ok, version}.",
     "uvicorn app.main:app sobe sem erro; GET /health retorna 200 com JSON esperado.",
     "S00-T03",
     "Teste integração: client.get('/health') -> status 200, body com status='ok'.",
     "Alta"),
    (1, "S01-T02", "Backend",
     "Settings via Pydantic-Settings",
     "Criar app/core/config.py com Settings(BaseSettings) lendo DATABASE_URL, JWT_SECRET, "
     "JWT_ALGORITHM, JWT_EXPIRE_MINUTES, ENVIRONMENT do .env.",
     "Settings carrega valores do .env; valores ausentes geram erro claro no boot.",
     "S01-T01",
     "Teste unit: instanciar Settings com .env de teste e validar campos.",
     "Alta"),
    (1, "S01-T03", "Banco",
     "Configurar SQLAlchemy e session",
     "Criar app/database/session.py com engine e SessionLocal; app/database/base.py com "
     "DeclarativeBase; dependency get_db() para FastAPI.",
     "Endpoint de teste consegue obter uma session válida; rollback automático em erro.",
     "S01-T02",
     "Teste integração: dependency get_db produz session que executa SELECT 1.",
     "Alta"),
    (1, "S01-T04", "Banco",
     "Configurar Alembic e primeira migration",
     "alembic init alembic; configurar env.py para usar Settings; criar migration vazia inicial.",
     "alembic upgrade head executa sem erro contra o Postgres do compose.",
     "S01-T03",
     "Teste manual: rodar upgrade e downgrade da migration inicial.",
     "Alta"),
    (1, "S01-T05", "Backend",
     "Logging estruturado e tratamento global de erros",
     "Configurar logging (JSON em prod, texto em dev). ExceptionHandler para HTTPException, "
     "ValidationError e Exception genérica, retornando payload padronizado "
     "{detail, code, request_id}.",
     "Erros não esperados retornam 500 com payload padronizado e log com request_id.",
     "S01-T01",
     "Testes: forçar 422 em endpoint de teste, forçar 500 em endpoint que levanta exceção, "
     "validar payload e log.",
     "Média"),

    # ---------------- Sprint 2 ----------------
    (2, "S02-T01", "Banco",
     "Model User e migration",
     "Criar app/users/models.py com User(id, name, email unique, password_hash, created_at). "
     "Gerar migration via alembic revision --autogenerate.",
     "Migration aplica e cria tabela users com índice único em email.",
     "S01-T04",
     "Teste: upgrade/downgrade da migration; insert duplicado falha por unique.",
     "Alta"),
    (2, "S02-T02", "Backend",
     "Utilitários de hash de senha",
     "Criar app/auth/security.py com hash_password(plain) e verify_password(plain, hashed) "
     "usando passlib[bcrypt].",
     "Funções determinísticas: verify retorna True/False corretamente.",
     "S00-T03",
     "Unit tests: hashing diferente para mesma senha (salt), verify correto e incorreto, "
     "verify de hash malformado não estoura.",
     "Alta"),
    (2, "S02-T03", "Backend",
     "Emissão e verificação de JWT",
     "Criar app/auth/jwt.py com create_access_token(user_id, expires) e decode_token(token) "
     "usando python-jose.",
     "Token gerado pode ser decodificado e contém sub=user_id; tokens expirados são rejeitados.",
     "S02-T02",
     "Unit tests: token válido decoda; token expirado levanta exceção; token assinado com "
     "outra chave levanta exceção.",
     "Alta"),
    (2, "S02-T04", "Backend",
     "Endpoint POST /auth/register",
     "Criar schema UserCreate (name, email, password). Validar email único, criar usuário "
     "com senha hasheada, retornar UserRead sem password.",
     "Registro de novo usuário retorna 201 com payload sem senha; email duplicado retorna 409.",
     "S02-T01, S02-T02",
     "Integration tests: cadastro feliz; email duplicado retorna 409; payload inválido retorna 422.",
     "Alta"),
    (2, "S02-T05", "Backend",
     "Endpoint POST /auth/login",
     "Receber email+password, validar, retornar access_token e token_type.",
     "Login com credenciais válidas retorna 200+JWT; credenciais inválidas retornam 401.",
     "S02-T03, S02-T04",
     "Integration tests: login feliz; senha errada -> 401; email inexistente -> 401; "
     "payload faltando campo -> 422.",
     "Alta"),
    (2, "S02-T06", "Backend",
     "Dependency get_current_user e GET /users/me",
     "Criar dependency que decoda JWT do header Authorization, carrega user do banco, "
     "retorna instância de User. Criar GET /users/me.",
     "/users/me com token válido retorna o usuário; sem token retorna 401; token inválido retorna 401.",
     "S02-T03, S02-T04",
     "Integration tests: /me sem header; /me com token aleatório; /me com token expirado; "
     "/me com token de usuário deletado.",
     "Alta"),
    (2, "S02-T07", "Teste",
     "Suite de testes de autenticação consolidada",
     "Revisar cobertura do módulo auth e garantir >= 90%. Adicionar testes que faltarem.",
     "pytest --cov=app.auth >= 90%.",
     "S02-T06",
     "Verificação de cobertura no CI.",
     "Média"),

    # ---------------- Sprint 3 ----------------
    (3, "S03-T01", "Banco",
     "Model Account e migration",
     "Criar Account(id, user_id FK, name, type enum, initial_balance, current_balance, "
     "is_active, created_at). Migration via autogenerate.",
     "Tabela accounts criada com FK e índice em user_id.",
     "S02-T01",
     "Teste: upgrade migration; insert com user_id inválido falha por FK.",
     "Alta"),
    (3, "S03-T02", "Backend",
     "Schemas Pydantic para Account",
     "AccountCreate, AccountUpdate, AccountRead. Validar type contra enum.",
     "Schemas validam corretamente; tipo desconhecido retorna 422.",
     "S03-T01",
     "Unit tests dos schemas.",
     "Média"),
    (3, "S03-T03", "Backend",
     "Repository e Service de Account",
     "AccountRepository com métodos CRUD filtrando por user_id. AccountService aplicando "
     "regras (current_balance = initial_balance na criação).",
     "Repository nunca expõe contas de outro usuário; service inicia current_balance corretamente.",
     "S03-T02",
     "Unit tests do service e repository com SQLite em memória ou Postgres de teste.",
     "Alta"),
    (3, "S03-T04", "Backend",
     "Routes /accounts (CRUD)",
     "GET (lista), POST (cria), GET/{id}, PATCH/{id}, DELETE/{id} — todos exigindo "
     "autenticação e isolados por usuário.",
     "Todos os endpoints respondem corretamente; usuário A nunca vê/edita conta do usuário B.",
     "S03-T03, S02-T06",
     "Integration tests: CRUD completo; isolamento entre usuários (criar com user A, tentar "
     "ler/editar/deletar com user B -> 404).",
     "Alta"),

    # ---------------- Sprint 4 ----------------
    (4, "S04-T01", "Banco",
     "Model Category e migration",
     "Category(id, user_id FK, name, type [receita/despesa], color, icon, is_default).",
     "Tabela categories criada com índice em user_id.",
     "S02-T01",
     "Teste de migration.",
     "Alta"),
    (4, "S04-T02", "Backend",
     "Seed de categorias padrão por usuário",
     "Função seed_default_categories(user_id) chamada após /auth/register, criando "
     "Alimentação, Transporte, Saúde, Educação, Moradia, Lazer, Assinaturas, Outros.",
     "Novo usuário tem 8 categorias padrão criadas com cores e ícones definidos.",
     "S04-T01, S02-T04",
     "Integration test: registrar usuário, listar categorias, esperar 8 itens com "
     "is_default=True.",
     "Alta"),
    (4, "S04-T03", "Backend",
     "Schemas, repository, service e routes /categories",
     "CRUD completo com isolamento por usuário. Impedir deleção de categoria padrão se houver "
     "transações associadas (após Sprint 5 essa regra ganhará força).",
     "CRUD funcional; categorias do usuário A invisíveis ao B.",
     "S04-T02",
     "Integration tests cobrindo CRUD, isolamento, criação com cor/ícone custom.",
     "Alta"),

    # ---------------- Sprint 5 ----------------
    (5, "S05-T01", "Banco",
     "Model Transaction e migration",
     "Transaction(id, user_id FK, account_id FK, category_id FK, type, amount, date, "
     "description, payment_method, is_recurring, created_at).",
     "Tabela transactions criada com FKs e índices em (user_id, date) e (account_id, date).",
     "S03-T01, S04-T01",
     "Teste de migration.",
     "Alta"),
    (5, "S05-T02", "Backend",
     "Schemas e validações de Transaction",
     "TransactionCreate exige account_id e category_id pertencentes ao usuário. "
     "amount > 0; type ∈ {receita, despesa}.",
     "Schemas e validações cruzadas funcionando.",
     "S05-T01",
     "Unit tests dos schemas; validação de account/category de outro usuário rejeitada.",
     "Alta"),
    (5, "S05-T03", "Backend",
     "Service: atualização de saldo (regra crítica)",
     "Ao criar/editar/deletar transação, recalcular ou ajustar current_balance da conta "
     "afetada. Receita soma, despesa subtrai. Transações entre contas (futuro) ficam fora.",
     "current_balance reflete exatamente a soma das transações; testes de regressão impedem "
     "que mudanças futuras quebrem essa regra.",
     "S05-T02, S03-T03",
     "TESTES OBRIGATÓRIOS: criar receita aumenta saldo; criar despesa diminui; editar valor "
     "ajusta delta; mudar tipo (receita->despesa) ajusta corretamente; deletar reverte; "
     "criar várias e validar soma final; concurrency simples (duas transações sequenciais).",
     "Alta"),
    (5, "S05-T04", "Backend",
     "Routes /transactions com filtros e paginação",
     "GET com query params: date_from, date_to, account_id, category_id, type, search (texto), "
     "page, page_size, order_by. CRUD completo.",
     "Filtros combinados funcionam; paginação retorna metadados (total, page, page_size).",
     "S05-T03",
     "Integration tests: cada filtro isolado; combinação de filtros; paginação; busca textual "
     "case-insensitive; isolamento entre usuários.",
     "Alta"),
    (5, "S05-T05", "Teste",
     "Bateria de regressão de saldo",
     "Conjunto de testes end-to-end que executa sequência realista (criar conta, várias "
     "transações, edições, deleções) e valida saldo final esperado.",
     "Sequência de operações resulta em saldo exato; teste serve de guarda permanente contra "
     "regressões em qualquer sprint futura.",
     "S05-T04",
     "Cenário longo e nomeado test_balance_regression_full_flow.",
     "Alta"),

    # ---------------- Sprint 6 ----------------
    (6, "S06-T01", "Banco",
     "Model Budget e migration",
     "Budget(id, user_id, category_id, month, year, limit_amount). Unique (user_id, "
     "category_id, month, year).",
     "Tabela criada; insert duplicado da mesma combinação falha.",
     "S04-T01",
     "Teste de migration e unique constraint.",
     "Alta"),
    (6, "S06-T02", "Backend",
     "Service: cálculo de uso e alertas",
     "Para um budget, calcular soma de despesas da categoria no mês/ano correspondente, "
     "retornar percent_used e status (ok, warning>80%, critical>100%).",
     "Cálculo correto considerando apenas despesas e apenas datas dentro do mês.",
     "S06-T01, S05-T03",
     "Unit tests: sem transações -> 0%; metade do limite -> 50%/ok; 85% -> warning; "
     "120% -> critical; transação de mês adjacente não conta; transação de receita não conta.",
     "Alta"),
    (6, "S06-T03", "Backend",
     "Routes /budgets",
     "GET (lista do mês corrente com status calculado), POST, PATCH. Comparativo mensal "
     "via query month/year.",
     "Endpoints respondem com payload incluindo percent_used e status.",
     "S06-T02",
     "Integration tests cobrindo mês corrente, mês passado, isolamento por usuário.",
     "Alta"),

    # ---------------- Sprint 7 ----------------
    (7, "S07-T01", "Banco",
     "Model Goal e migration",
     "Goal(id, user_id, name, target_amount, current_amount, deadline, status).",
     "Tabela criada.",
     "S02-T01",
     "Teste de migration.",
     "Alta"),
    (7, "S07-T02", "Backend",
     "Service e routes /goals",
     "CRUD completo. PATCH para atualizar current_amount; recalcular status como concluído "
     "quando current >= target.",
     "Atualização para current >= target marca status concluído automaticamente; reverter "
     "current abaixo volta para em_andamento.",
     "S07-T01",
     "Unit + integration tests: criação; progresso parcial; transição automática para concluído; "
     "reversão; deadline ultrapassada não muda status sozinha.",
     "Alta"),

    # ---------------- Sprint 8 ----------------
    (8, "S08-T01", "Backend",
     "GET /dashboard/monthly-summary",
     "Retorna {receita_total, despesa_total, saldo, contas: [...]} para mês/ano dado.",
     "Soma correta; mês sem transações retorna zeros; respeita usuário autenticado.",
     "S05-T03",
     "Integration tests com cenários: mês vazio; mês com receitas e despesas; várias contas.",
     "Alta"),
    (8, "S08-T02", "Backend",
     "GET /dashboard/category-breakdown",
     "Retorna lista de {category_id, name, color, total} ordenado desc por total, "
     "filtrado por mês/ano.",
     "Lista correta; categorias sem transações não aparecem (ou aparecem com 0 conforme decisão).",
     "S08-T01",
     "Integration tests cobrindo várias categorias e mês vazio.",
     "Média"),
    (8, "S08-T03", "Backend",
     "GET /dashboard/cashflow",
     "Retorna série temporal {date, receita, despesa, saldo_acumulado} para período dado.",
     "Série correta e ordenada por data.",
     "S08-T01",
     "Integration tests com período curto e longo.",
     "Média"),

    # ---------------- Sprint 9 ----------------
    (9, "S09-T01", "Backend",
     "Endpoint POST /imports/csv",
     "Aceitar multipart/form-data com arquivo CSV e account_id alvo. Validar tamanho máx, "
     "extensão e MIME.",
     "Upload válido aceito; arquivo inválido rejeitado com 400.",
     "S05-T03",
     "Integration tests: CSV válido, CSV vazio, arquivo não-CSV, arquivo gigante.",
     "Alta"),
    (9, "S09-T02", "Backend",
     "Parser CSV e categorização automática",
     "Parser robusto suportando os formatos mais comuns (data, descrição, valor). Aplicar "
     "regras: UBER->Transporte, IFOOD->Alimentação, NETFLIX->Assinaturas, fallback Outros.",
     "Linhas válidas viram transações; linhas inválidas viram lista de erros no payload de resposta.",
     "S09-T01",
     "Unit tests do parser com CSVs sintéticos; tests das regras de categorização.",
     "Alta"),
    (9, "S09-T03", "Backend",
     "Criação em batch e atualização de saldos",
     "Criar transações em transação única; reverter tudo em caso de erro; atualizar "
     "current_balance da conta uma única vez ao final.",
     "Saldo da conta reflete exatamente a soma das transações importadas; em erro nada é gravado.",
     "S09-T02, S05-T03",
     "Integration tests: import feliz; import com erro no meio (rollback); validar saldo final.",
     "Alta"),

    # ---------------- Sprint 10 ----------------
    (10, "S10-T01", "Backend",
     "CORS configurado por ambiente",
     "Liberar origens via lista no Settings; em dev permitir localhost; em prod só os domínios "
     "do app/painel.",
     "Requisições de origem permitida passam; outras são bloqueadas.",
     "S01-T05",
     "Testes manuais e integração com OPTIONS request.",
     "Média"),
    (10, "S10-T02", "Backend",
     "Rate limiting básico",
     "Aplicar limite por IP nos endpoints públicos (/auth/login, /auth/register).",
     "Após N tentativas em janela X, retornar 429.",
     "S02-T05",
     "Integration tests: estourar limite e validar 429.",
     "Média"),
    (10, "S10-T03", "Backend",
     "Refresh token",
     "Adicionar refresh_token na resposta de login; endpoint POST /auth/refresh.",
     "Refresh válido emite novo access; refresh inválido/expirado retorna 401.",
     "S02-T05",
     "Integration tests cobrindo fluxo de refresh.",
     "Média"),
    (10, "S10-T04", "Documentação",
     "OpenAPI completo",
     "Garantir summary, description, response_model e exemplos em todos os endpoints. "
     "Tags organizadas.",
     "Swagger UI navegável e legível; cliente externo consegue gerar SDK a partir do schema.",
     "Todos endpoints anteriores",
     "Revisão visual em /docs.",
     "Média"),
    (10, "S10-T05", "Teste",
     "Cobertura mínima 80% do backend",
     "Rodar pytest --cov e completar testes nos módulos abaixo do alvo.",
     "Relatório de coverage >= 80% global e >= 70% por módulo.",
     "Todos anteriores",
     "Relatório automático no CI falhando abaixo do alvo.",
     "Alta"),

    # ---------------- Sprint 11 ----------------
    (11, "S11-T01", "Mobile",
     "Inicializar projeto Flutter",
     "flutter create mobile; configurar análise (analysis_options.yaml estrito), "
     "dependências Riverpod, go_router, Dio, fl_chart, flutter_secure_storage, "
     "flutter_test, mocktail.",
     "flutter pub get sem erros; flutter analyze limpo; flutter test rodando exemplo.",
     "S00-T01",
     "flutter analyze e flutter test no CI mobile.",
     "Alta"),
    (11, "S11-T02", "Mobile",
     "Camada de rede com Dio e interceptor JWT",
     "Configurar Dio com baseUrl do .env, interceptor que injeta Authorization, "
     "interceptor que detecta 401 e força logout.",
     "Requisições autenticadas anexam header; 401 limpa storage e navega para Login.",
     "S11-T01",
     "Widget/unit tests com Dio mockado.",
     "Alta"),
    (11, "S11-T03", "Mobile",
     "Telas Splash, Login e Cadastro",
     "Navegação via go_router; estado via Riverpod; validação de formulário; mensagens de erro "
     "claras; salvar token no flutter_secure_storage após login.",
     "Fluxo Splash -> Login -> Dashboard quando há token; Splash -> Login quando não há.",
     "S11-T02, S02-T05",
     "Widget tests dos formulários; teste de navegação após login bem-sucedido (mock do Dio).",
     "Alta"),

    # ---------------- Sprint 12 ----------------
    (12, "S12-T01", "Mobile",
     "Tela Dashboard com cards e fl_chart",
     "Cards de receita/despesa/saldo do mês; gráfico de pizza por categoria; gráfico de linha "
     "do cashflow. Pull-to-refresh.",
     "Dados carregados do backend; estados vazio, carregando e erro tratados.",
     "S08-T03, S11-T03",
     "Widget tests cobrindo os 3 estados.",
     "Alta"),
    (12, "S12-T02", "Mobile",
     "Tela Lista de Transações com filtros",
     "Listagem paginada (scroll infinito); filtros por data, categoria, conta, tipo; busca textual.",
     "Filtros aplicados refletem no payload; estados vazio/erro tratados.",
     "S05-T04, S11-T03",
     "Widget tests dos filtros e scroll.",
     "Alta"),
    (12, "S12-T03", "Mobile",
     "Telas Nova e Editar Transação",
     "Formulário com selects de conta e categoria, date picker, validação de valor.",
     "Criar e editar funcionam; lista é atualizada após operação.",
     "S12-T02",
     "Widget tests dos formulários.",
     "Alta"),

    # ---------------- Sprint 13 ----------------
    (13, "S13-T01", "Mobile",
     "Tela Categorias",
     "CRUD visual de categorias com seletor de cor e ícone.",
     "CRUD funciona contra o backend; alterações refletem em transações.",
     "S04-T03, S11-T03",
     "Widget tests.",
     "Média"),
    (13, "S13-T02", "Mobile",
     "Tela Orçamento com alertas visuais",
     "Listar orçamentos do mês corrente com barra de progresso colorida (verde/amarelo/vermelho) "
     "segundo o status.",
     "Cores e percentuais batem com o backend; criar/editar orçamento funciona.",
     "S06-T03, S11-T03",
     "Widget tests para os três estados de status.",
     "Alta"),
    (13, "S13-T03", "Mobile",
     "Tela Metas com progresso",
     "Listar metas com barra de progresso; criar e atualizar; marcação automática quando concluída.",
     "Progresso e status corretos; transição visual quando atinge 100%.",
     "S07-T02, S11-T03",
     "Widget tests cobrindo progresso parcial e total.",
     "Média"),

    # ---------------- Sprint 14 ----------------
    (14, "S14-T01", "Mobile",
     "Tela Perfil",
     "Mostrar dados do usuário, opção de logout.",
     "Logout limpa token e volta para Login.",
     "S11-T03",
     "Widget test do logout.",
     "Baixa"),
    (14, "S14-T02", "Mobile",
     "Tela Configurações",
     "Tema (claro/escuro/sistema), preferências básicas.",
     "Tema persiste entre sessões.",
     "S14-T01",
     "Widget tests.",
     "Baixa"),
    (14, "S14-T03", "Mobile",
     "Upload de CSV",
     "file_picker para selecionar CSV, enviar para /imports/csv, exibir resumo (sucesso/erros).",
     "Fluxo funciona ponta a ponta; erros do backend exibidos amigavelmente.",
     "S09-T03, S11-T02",
     "Widget tests com mock do upload.",
     "Média"),

    # ---------------- Sprint 15 ----------------
    (15, "S15-T01", "Frontend",
     "Setup do painel Streamlit",
     "panel/app.py com Streamlit; tela de login consumindo /auth/login; sessão guardada em "
     "st.session_state.",
     "Login funciona; logout limpa estado.",
     "S02-T05",
     "Testes manuais documentados; smoke test com Selenium opcional.",
     "Média"),
    (15, "S15-T02", "Frontend",
     "Relatórios mensais e por categoria",
     "Páginas Streamlit consumindo /dashboard/*; gráficos Plotly/Altair; filtros de período.",
     "Relatórios refletem dados do backend; filtros funcionam.",
     "S08-T03, S15-T01",
     "Testes manuais documentados.",
     "Média"),
    (15, "S15-T03", "Frontend",
     "Exportação básica (CSV/Excel)",
     "Botão para exportar transações filtradas em CSV/XLSX.",
     "Arquivo gerado abre corretamente.",
     "S15-T02",
     "Teste manual com arquivo gerado.",
     "Baixa"),

    # ---------------- Sprint 16 ----------------
    (16, "S16-T01", "Deploy",
     "Dockerfile de produção do backend",
     "Imagem multi-stage, usuário não-root, healthcheck, gunicorn/uvicorn workers.",
     "docker build conclui; container roda /health verde.",
     "S10-T05",
     "Build local + run com healthcheck.",
     "Alta"),
    (16, "S16-T02", "Deploy",
     "Provisionar banco gerenciado (Supabase)",
     "Criar projeto Supabase; aplicar migrations Alembic via DATABASE_URL remoto; ajustar "
     "pool size.",
     "Migrations aplicadas com sucesso no banco gerenciado.",
     "S16-T01",
     "Smoke test: SELECT 1 e criação de usuário via API contra o banco gerenciado.",
     "Alta"),
    (16, "S16-T03", "Deploy",
     "Deploy do backend (Railway ou Render)",
     "Configurar secrets, deploy automático a partir do main; domínio HTTPS.",
     "/health responde 200 no domínio público.",
     "S16-T02",
     "Smoke tests pós-deploy: /health, /auth/register, /auth/login.",
     "Alta"),
    (16, "S16-T04", "Mobile",
     "Build do APK Android",
     "flutter build apk --release apontando para o backend de produção; assinatura básica.",
     "APK instalável; login contra backend de produção funciona.",
     "S16-T03, S14-T03",
     "Teste manual em device/emulador Android.",
     "Média"),
    (16, "S16-T05", "Documentação",
     "Documentação final e README do projeto",
     "Consolidar README com arquitetura, decisões, links de deploy, instruções de uso, "
     "limitações e roadmap futuro.",
     "Documentação navegável e suficiente para um terceiro entender e operar o projeto.",
     "S16-T03, S16-T04",
     "Revisão de conteúdo.",
     "Média"),
    (16, "S16-T06", "Teste",
     "Smoke tests de produção",
     "Script (pytest ou bash) que valida endpoints críticos em produção sem destruir dados.",
     "Script roda verde no ambiente prod e pode ser executado periodicamente.",
     "S16-T03",
     "Execução pós-deploy.",
     "Alta"),
]


# ============================================================
# Plano de Testes (resumo por feature)
# ============================================================

TESTES = [
    # (sprint, modulo, tipo_teste, descricao, sprint_alvo, status_inicial)
    (1, "core", "Integração", "GET /health responde 200", 1, "Pendente"),
    (1, "core", "Unitário", "Settings carrega .env de teste", 1, "Pendente"),
    (1, "database", "Integração", "Migration inicial up/down", 1, "Pendente"),
    (1, "errors", "Integração", "ExceptionHandler retorna payload padronizado", 1, "Pendente"),

    (2, "auth", "Unitário", "hash/verify de senha", 2, "Pendente"),
    (2, "auth", "Unitário", "JWT issue/decode/expirado/chave errada", 2, "Pendente"),
    (2, "auth", "Integração", "Register feliz / email duplicado / inválido", 2, "Pendente"),
    (2, "auth", "Integração", "Login feliz / senha errada / email inexistente", 2, "Pendente"),
    (2, "auth", "Integração", "/users/me sem token / token inválido / expirado / usuário deletado", 2, "Pendente"),

    (3, "accounts", "Integração", "CRUD completo de Account", 3, "Pendente"),
    (3, "accounts", "Integração", "Isolamento de Account entre usuários", 3, "Pendente"),

    (4, "categories", "Integração", "Seed das 8 categorias padrão no registro", 4, "Pendente"),
    (4, "categories", "Integração", "CRUD de Category com isolamento", 4, "Pendente"),

    (5, "transactions", "Unitário", "Receita aumenta saldo / Despesa diminui", 5, "Pendente"),
    (5, "transactions", "Unitário", "Edição ajusta delta corretamente", 5, "Pendente"),
    (5, "transactions", "Unitário", "Mudança de tipo (receita<->despesa) ajusta saldo", 5, "Pendente"),
    (5, "transactions", "Unitário", "Deleção reverte saldo", 5, "Pendente"),
    (5, "transactions", "Integração", "Filtros, paginação, busca textual", 5, "Pendente"),
    (5, "transactions", "Regressão", "Cenário longo end-to-end de saldo (test_balance_regression_full_flow)", 5, "Pendente"),

    (6, "budgets", "Unitário", "Cálculo de percent_used e status (ok/warning/critical)", 6, "Pendente"),
    (6, "budgets", "Integração", "Comparativo mensal e isolamento", 6, "Pendente"),

    (7, "goals", "Integração", "Transição automática para concluída e reversão", 7, "Pendente"),

    (8, "dashboard", "Integração", "monthly-summary vazio e populado", 8, "Pendente"),
    (8, "dashboard", "Integração", "category-breakdown e cashflow", 8, "Pendente"),

    (9, "imports", "Unitário", "Parser CSV com formatos válidos e inválidos", 9, "Pendente"),
    (9, "imports", "Unitário", "Regras de categorização (UBER/IFOOD/NETFLIX/fallback)", 9, "Pendente"),
    (9, "imports", "Integração", "Import feliz atualiza saldo; erro causa rollback total", 9, "Pendente"),

    (10, "auth", "Integração", "Rate limiting em /auth/login retorna 429", 10, "Pendente"),
    (10, "auth", "Integração", "Refresh token feliz / inválido / expirado", 10, "Pendente"),
    (10, "global", "Cobertura", "pytest --cov >= 80% global", 10, "Pendente"),

    (11, "mobile", "Widget", "Splash navega para Login ou Dashboard conforme token", 11, "Pendente"),
    (11, "mobile", "Widget", "Login com mock do Dio (sucesso e erro)", 11, "Pendente"),
    (11, "mobile", "Widget", "Cadastro com mock do Dio", 11, "Pendente"),

    (12, "mobile", "Widget", "Dashboard nos estados vazio/carregando/erro", 12, "Pendente"),
    (12, "mobile", "Widget", "Lista de transações com filtros e scroll", 12, "Pendente"),
    (12, "mobile", "Widget", "Formulários de criar/editar transação", 12, "Pendente"),

    (13, "mobile", "Widget", "Orçamento exibe corretamente nos 3 status", 13, "Pendente"),
    (13, "mobile", "Widget", "Metas em progresso parcial e total", 13, "Pendente"),
    (13, "mobile", "Widget", "Categorias CRUD", 13, "Pendente"),

    (14, "mobile", "Widget", "Logout limpa token e volta para Login", 14, "Pendente"),
    (14, "mobile", "Widget", "Upload CSV exibe sucesso/erro", 14, "Pendente"),

    (15, "panel", "Manual", "Login no painel Streamlit", 15, "Pendente"),
    (15, "panel", "Manual", "Relatórios e filtros do painel", 15, "Pendente"),

    (16, "deploy", "Smoke", "/health em produção", 16, "Pendente"),
    (16, "deploy", "Smoke", "Register e login em produção", 16, "Pendente"),
    (16, "deploy", "Manual", "APK instala e loga contra produção", 16, "Pendente"),
]


# ============================================================
# Construção da planilha
# ============================================================

def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER


def style_body(cell):
    cell.alignment = WRAP
    cell.border = BORDER


def set_columns(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


# ---- Aba Instruções ----

def build_instructions(wb):
    ws = wb.create_sheet("Instruções")
    set_columns(ws, [110])

    blocks = [
        ("OrçaFácil — Documentação Verdade das Sprints", "title"),
        ("Esta planilha é a fonte única de verdade para a execução do projeto. O chat do "
         "Claude Code que vai trabalhar nas tarefas deve usá-la como roteiro, registrar "
         "evolução nela e nunca tomar decisões de escopo fora dela.", "note"),
        ("", None),
        ("Como o Claude Code deve usar esta planilha", "section"),
        ("1. Sempre comece a sessão lendo a aba 'Visão Geral' para entender o estado macro.", "body"),
        ("2. Em seguida, abra a aba 'Backlog' e localize a próxima tarefa com status "
         "'Pendente' respeitando a coluna 'Dependências' (não pular ordem).", "body"),
        ("3. Marque a tarefa como 'Em andamento' e preencha 'Data Início'.", "body"),
        ("4. Implemente a tarefa seguindo exatamente o que está em 'Descrição' e "
         "'Critérios de Aceitação'. Não adicionar features extras.", "body"),
        ("5. Escrever os testes listados em 'Testes Obrigatórios' ANTES de marcar como "
         "concluída. Sem testes verdes a tarefa não é concluída.", "body"),
        ("6. Rodar a suíte completa de testes do módulo afetado para garantir que nada "
         "regrediu antes de concluir.", "body"),
        ("7. Mudar status para 'Em teste' quando o código estiver pronto e os testes "
         "estiverem rodando; mudar para 'Concluída' quando tudo passar.", "body"),
        ("8. Preencher 'Data Fim' e qualquer 'Observação' relevante (decisões tomadas, "
         "arquivos criados, comandos úteis para rodar).", "body"),
        ("9. Se houver dúvida que mude o escopo ou comportamento esperado da tarefa, "
         "marcar 'Impedimento? = Sim', escrever a dúvida e PARAR. Perguntar para o "
         "humano via mensagem. Nunca decidir sozinho.", "body"),
        ("", None),
        ("Regras inegociáveis", "section"),
        ("• Toda tarefa que toca regra de negócio (saldo, alerta de orçamento, status "
         "de meta, import CSV) DEVE ter testes que travam a regressão.", "body"),
        ("• A bateria de regressão de saldo (S05-T05) deve continuar verde em todas as "
         "sprints seguintes. Se quebrar, parar e investigar.", "body"),
        ("• Cobertura de testes do backend deve permanecer ≥ 80% a partir da Sprint 10.", "body"),
        ("• Nunca commitar segredos. Sempre usar .env e .env.example.", "body"),
        ("• Nunca apagar dados em produção sem autorização explícita do humano.", "body"),
        ("• Mudanças de schema sempre via migration Alembic — nunca direto no banco.", "body"),
        ("• Para qualquer decisão técnica não trivial, registrar na aba 'Decisões'.", "body"),
        ("", None),
        ("Sobre a aba 'Plano de Testes'", "section"),
        ("Lista de checks de teste que precisam existir ao final do projeto. Marcar como "
         "'Concluída' assim que o teste correspondente estiver no repositório e verde no CI.", "body"),
        ("", None),
        ("Sobre a aba 'Impedimentos'", "section"),
        ("Registrar QUALQUER bloqueio antes de prosseguir. Cada linha tem ID, sprint/tarefa "
         "relacionada, descrição do problema, o que tentou e a pergunta para o humano.", "body"),
        ("", None),
        ("Sobre a aba 'Decisões'", "section"),
        ("Registrar decisões técnicas não óbvias: por que escolheu biblioteca X, por que "
         "estrutura Y, etc. Garante que o próximo Claude entenda o porquê.", "body"),
        ("", None),
        ("Convenções", "section"),
        ("• IDs de tarefa no formato S##-T## (ex.: S05-T03).", "body"),
        ("• Status válidos: Pendente, Em andamento, Bloqueada, Em teste, Concluída, "
         "Cancelada (dropdown na coluna Status).", "body"),
        ("• Datas no formato AAAA-MM-DD.", "body"),
        ("• Observações curtas e úteis, sem narrativa longa.", "body"),
    ]

    row = 1
    for text, kind in blocks:
        cell = ws.cell(row=row, column=1, value=text)
        if kind == "title":
            cell.font = TITLE_FONT
            ws.row_dimensions[row].height = 28
        elif kind == "section":
            cell.font = SECTION_FONT
            ws.row_dimensions[row].height = 22
        elif kind == "note":
            cell.font = NOTE_FONT
            cell.alignment = WRAP
            ws.row_dimensions[row].height = 45
        else:
            cell.alignment = WRAP
            ws.row_dimensions[row].height = 30
        row += 1


# ---- Aba Visão Geral ----

def build_overview(wb):
    ws = wb.create_sheet("Visão Geral")
    headers = ["Sprint", "Nome", "Objetivo", "Entregáveis", "Status", "Data Início", "Data Fim", "Observações"]
    widths = [8, 50, 60, 60, 16, 14, 14, 50]
    set_columns(ws, widths)

    for col, h in enumerate(headers, start=1):
        style_header(ws.cell(row=1, column=col, value=h))

    for i, s in enumerate(SPRINTS, start=2):
        ws.cell(row=i, column=1, value=s["numero"])
        ws.cell(row=i, column=2, value=s["nome"])
        ws.cell(row=i, column=3, value=s["objetivo"])
        ws.cell(row=i, column=4, value=s["entregaveis"])
        ws.cell(row=i, column=5, value="Pendente")
        ws.cell(row=i, column=6, value="")
        ws.cell(row=i, column=7, value="")
        ws.cell(row=i, column=8, value="")
        for c in range(1, 9):
            cell = ws.cell(row=i, column=c)
            style_body(cell)
            if c == 1:
                cell.fill = SPRINT_FILL

        ws.row_dimensions[i].height = 75

    last_row = 1 + len(SPRINTS)
    dv = DataValidation(type="list", formula1=f'"{",".join(STATUS_OPTIONS)}"', allow_blank=True)
    dv.add(f"E2:E{last_row}")
    ws.add_data_validation(dv)

    ws.freeze_panes = "A2"


# ---- Aba Backlog ----

def build_backlog(wb):
    ws = wb.create_sheet("Backlog")
    headers = [
        "Sprint", "ID", "Tipo", "Prioridade", "Tarefa", "Descrição",
        "Critérios de Aceitação", "Dependências", "Testes Obrigatórios",
        "Status", "Data Início", "Data Fim", "Observações",
        "Impedimento?", "Detalhe do Impedimento",
    ]
    widths = [8, 10, 14, 11, 38, 60, 55, 18, 55, 14, 13, 13, 40, 13, 40]
    set_columns(ws, widths)

    for col, h in enumerate(headers, start=1):
        style_header(ws.cell(row=1, column=col, value=h))

    for i, t in enumerate(TAREFAS, start=2):
        sprint, tid, tipo, titulo, desc, criterios, deps, testes, prio = t
        ws.cell(row=i, column=1, value=sprint)
        ws.cell(row=i, column=2, value=tid)
        ws.cell(row=i, column=3, value=tipo)
        ws.cell(row=i, column=4, value=prio)
        ws.cell(row=i, column=5, value=titulo)
        ws.cell(row=i, column=6, value=desc)
        ws.cell(row=i, column=7, value=criterios)
        ws.cell(row=i, column=8, value=deps)
        ws.cell(row=i, column=9, value=testes)
        ws.cell(row=i, column=10, value="Pendente")
        ws.cell(row=i, column=11, value="")
        ws.cell(row=i, column=12, value="")
        ws.cell(row=i, column=13, value="")
        ws.cell(row=i, column=14, value="Não")
        ws.cell(row=i, column=15, value="")

        for c in range(1, 16):
            style_body(ws.cell(row=i, column=c))

        ws.row_dimensions[i].height = 110

    last_row = 1 + len(TAREFAS)

    dv_status = DataValidation(type="list", formula1=f'"{",".join(STATUS_OPTIONS)}"', allow_blank=True)
    dv_status.add(f"J2:J{last_row}")
    ws.add_data_validation(dv_status)

    dv_prio = DataValidation(type="list", formula1=f'"{",".join(PRIORIDADE_OPTIONS)}"', allow_blank=True)
    dv_prio.add(f"D2:D{last_row}")
    ws.add_data_validation(dv_prio)

    dv_tipo = DataValidation(type="list", formula1=f'"{",".join(TIPO_OPTIONS)}"', allow_blank=True)
    dv_tipo.add(f"C2:C{last_row}")
    ws.add_data_validation(dv_tipo)

    dv_imp = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True)
    dv_imp.add(f"N2:N{last_row}")
    ws.add_data_validation(dv_imp)

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:O{last_row}"


# ---- Aba Plano de Testes ----

def build_tests(wb):
    ws = wb.create_sheet("Plano de Testes")
    headers = ["Sprint Alvo", "Módulo", "Tipo de Teste", "Descrição", "Arquivo de Teste", "Status", "Última Execução", "Observações"]
    widths = [12, 16, 14, 70, 40, 14, 18, 40]
    set_columns(ws, widths)

    for col, h in enumerate(headers, start=1):
        style_header(ws.cell(row=1, column=col, value=h))

    for i, t in enumerate(TESTES, start=2):
        sprint, modulo, tipo, desc, sprint_alvo, status = t
        ws.cell(row=i, column=1, value=sprint_alvo)
        ws.cell(row=i, column=2, value=modulo)
        ws.cell(row=i, column=3, value=tipo)
        ws.cell(row=i, column=4, value=desc)
        ws.cell(row=i, column=5, value="")
        ws.cell(row=i, column=6, value=status)
        ws.cell(row=i, column=7, value="")
        ws.cell(row=i, column=8, value="")
        for c in range(1, 9):
            style_body(ws.cell(row=i, column=c))
        ws.row_dimensions[i].height = 35

    last_row = 1 + len(TESTES)
    dv = DataValidation(type="list", formula1=f'"{",".join(STATUS_OPTIONS)}"', allow_blank=True)
    dv.add(f"F2:F{last_row}")
    ws.add_data_validation(dv)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{last_row}"


# ---- Aba Impedimentos ----

def build_blockers(wb):
    ws = wb.create_sheet("Impedimentos")
    headers = ["ID", "Data", "Sprint", "Tarefa Relacionada", "Descrição", "O que tentou", "Pergunta para o humano", "Status", "Resposta / Decisão"]
    widths = [8, 14, 8, 14, 50, 50, 50, 16, 50]
    set_columns(ws, widths)

    for col, h in enumerate(headers, start=1):
        style_header(ws.cell(row=1, column=col, value=h))

    for i in range(2, 22):
        for c in range(1, 10):
            style_body(ws.cell(row=i, column=c))
        ws.row_dimensions[i].height = 35

    dv = DataValidation(type="list", formula1='"Aberto,Aguardando humano,Resolvido,Descartado"', allow_blank=True)
    dv.add("H2:H200")
    ws.add_data_validation(dv)

    ws.freeze_panes = "A2"


# ---- Aba Decisões ----

def build_decisions(wb):
    ws = wb.create_sheet("Decisões")
    headers = ["ID", "Data", "Sprint", "Contexto", "Decisão", "Alternativas consideradas", "Justificativa", "Impacto / Notas"]
    widths = [8, 14, 8, 50, 50, 50, 50, 40]
    set_columns(ws, widths)

    for col, h in enumerate(headers, start=1):
        style_header(ws.cell(row=1, column=col, value=h))

    for i in range(2, 22):
        for c in range(1, 9):
            style_body(ws.cell(row=i, column=c))
        ws.row_dimensions[i].height = 35

    ws.freeze_panes = "A2"


# ============================================================
# Build
# ============================================================

def main():
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    build_instructions(wb)
    build_overview(wb)
    build_backlog(wb)
    build_tests(wb)
    build_blockers(wb)
    build_decisions(wb)

    out_path = r"d:\Projetos\OrcamentoFamiliar\orcafacil_planejamento_sprints.xlsx"
    wb.save(out_path)
    print(f"Planilha gerada: {out_path}")
    print(f"Total de sprints: {len(SPRINTS)}")
    print(f"Total de tarefas: {len(TAREFAS)}")
    print(f"Total de itens de teste: {len(TESTES)}")


if __name__ == "__main__":
    main()
